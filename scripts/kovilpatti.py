import os
import textwrap
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# ==============================================================================
# STEP 1: DEFINE SYSTEM CONFIGURATION (EXACT 21 CANDIDATES)
# ==============================================================================
party_config = {
    'Dravida Munnetra Kazhagam': {'label': 'DMK', 'color': '#D32F2F'},                  
    'All India Anna Dravida Munnetra Kazhagam': {'label': 'AIADMK', 'color': '#2E7D32'}, 
    'Naam Tamilar Katchi': {'label': 'NTK', 'color': '#FBC02D'},                         
    'Tamilaga Vettri Kazhagam': {'label': 'TVK', 'color': '#7B1E1E'},                    
    'Bahujan Samaj Party': {'label': 'BSP', 'color': '#1F77B4'},                         
    'Naam Indiar Party': {'label': 'NIP', 'color': '#9E9E9E'},
    'All India Puratchi Thalaivar Makkal Munnettra Kazhagam': {'label': 'AIPMMK', 'color': '#455A64'},
    'Thamizh Perarasu Katchi': {'label': 'TPK', 'color': '#FF8C00'},
    'Puthiya Makkal Tamil Desam Katchi': {'label': 'PMTDK', 'color': '#00897B'},
    'Independent': {'label': 'IND1', 'color': '#BDBDBD'},
    'Independent.1': {'label': 'IND2', 'color': '#9E9E9E'},
    'Independent.2': {'label': 'IND3', 'color': '#757575'},
    'Independent.3': {'label': 'IND4', 'color': '#616161'},
    'Independent.4': {'label': 'IND5', 'color': '#424242'},
    'Independent.5': {'label': 'IND6', 'color': '#A6A6A6'},
    'Independent.6': {'label': 'IND7', 'color': '#7F7F7F'},
    'Independent.7': {'label': 'IND8', 'color': '#D9D9D9'},
    'Independent.8': {'label': 'IND9', 'color': '#595959'},
    'Independent.9': {'label': 'IND10', 'color': '#CCCCCC'},
    'Independent.10': {'label': 'IND11', 'color': '#333333'},
    'Independent.11': {'label': 'IND12', 'color': '#262626'}
}

party_columns = list(party_config.keys())
independent_cols = [c for c in party_columns if 'Independent' in c]

output_dir = 'expanded_polling_charts'
os.makedirs(output_dir, exist_ok=True)
excel_filename = 'Constituency_Expanded_Intelligence_Report.xlsx'

# ==============================================================================
# DATA SANITIZATION LAYER: ALPHANUMERIC STRING MANAGEMENT
# ==============================================================================
print("Standardising column layouts and running alphanumeric configuration...")

df.columns = df.columns.str.replace(r'\r', '', regex=True)
df.columns = df.columns.str.replace(r'\s+', ' ', regex=True).str.strip()

station_col = 'Serial No. Of Polling Station'
locality_col = 'Locality'
area_col = 'Polling Area' if 'Polling Area' in df.columns else 'Polling  Area'
valid_votes_col = 'Total of Valid\nVotes' if 'Total of Valid\nVotes' in df.columns else 'Total of Valid Votes'

# Clear calculation leftovers from previous notebook cell runs
columns_to_drop = [f'{col}_Rank' for col in party_columns] + [
    'Winner_Party', 'Winner_Votes', 'Runner_Up_Votes', 'Margin_Of_Victory', 
    'Runner_Up_Party', 'Margin_Percentage', 'Total_Independent_Votes', 'Unnamed: 0',
    'Total_Valid_Votes_Clean', 'Sort_Key', 'Sort_Suffix'
]
df = df.drop(columns=[c for c in columns_to_drop if c in df.columns], errors='ignore')

# Retain numbers as clean string tags to preserve alpha suffixes (e.g., '289A')
df[station_col] = df[station_col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
df['Total_Valid_Votes_Clean'] = pd.to_numeric(df[valid_votes_col], errors='coerce').fillna(0).astype(int)

# Create a composite sorting index so that numerical sequences separate suffixes properly
df['Sort_Key'] = df[station_col].str.extract(r'(\d+)').astype(float).fillna(0)
df['Sort_Suffix'] = df[station_col].str.extract(r'([A-Za-z]+)').fillna('')

# Apply the alphanumeric sorting index parameters
df = df.sort_values(by=['Sort_Key', 'Sort_Suffix'], ascending=[True, True]).reset_index(drop=True)
df = df.drop(columns=['Sort_Key', 'Sort_Suffix'])

# ==============================================================================
# STEP 2: PRIMARY VOTING CALCULATIONS (WINNER, MARGIN, AND RANKINGS)
# ==============================================================================
print("Running complete ranking matrix calculations...")

df['Winner_Party'] = df[party_columns].idxmax(axis=1)
df['Winner_Votes'] = df[party_columns].max(axis=1)

sorted_votes = np.sort(df[party_columns].values, axis=1)
df['Runner_Up_Votes'] = sorted_votes[:, -2]
df['Margin_Of_Victory'] = df['Winner_Votes'] - df['Runner_Up_Votes']

sorted_indices = np.argsort(df[party_columns].values, axis=1)
df['Runner_Up_Party'] = [party_columns[idx[-2]] for idx in sorted_indices]

ranks_df = df[party_columns].rank(axis=1, ascending=False, method='min').astype(int)
ranks_df = ranks_df.rename(columns={col: f'{col}_Rank' for col in party_columns})
df = pd.concat([df, ranks_df], axis=1)

# ==============================================================================
# STEP 3: STRATEGIC ANCILLARY METRICS (SWINGS, DOMINANCE, SPLITTING)
# ==============================================================================
print("Isolating battleground regions and independent impact scales...")

df['Margin_Percentage'] = (df['Margin_Of_Victory'] / df['Total_Valid_Votes_Clean']) * 100

critical_swings = df[df['Margin_Percentage'] <= 5.0][
    [station_col, locality_col, area_col, 'Winner_Party', 'Runner_Up_Party', 'Margin_Of_Victory', 'Margin_Percentage']
] 

locality_counts = df.groupby([locality_col, 'Winner_Party']).size().unstack(fill_value=0)
locality_dominance_summary = pd.DataFrame({
    'Total_Booths_In_Locality': df.groupby(locality_col).size(),
    'Dominant_Party': locality_counts.idxmax(axis=1),
    'Dominant_Party_Booths_Won': locality_counts.max(axis=1)
}).reset_index()

df['Total_Independent_Votes'] = df[independent_cols].sum(axis=1)
independent_vote_splitting = df.groupby(locality_col).agg(
    Total_Valid_Votes_In_Locality=('Total_Valid_Votes_Clean', 'sum'),
    Total_Independent_Votes_In_Locality=('Total_Independent_Votes', 'sum')
).reset_index()
independent_vote_splitting['Independent_Vote_Share_%'] = (
    independent_vote_splitting['Total_Independent_Votes_In_Locality'] / 
    independent_vote_splitting['Total_Valid_Votes_In_Locality']
) * 100
independent_vote_splitting = independent_vote_splitting.sort_values(by='Independent_Vote_Share_%', ascending=False)

# ==============================================================================
# STEP 4: GENERATE MACRO CONSTITUENCY DASHBOARD CHART
# ==============================================================================
print("Compiling constituency macro dashboard visual layer...")
total_valid_constituency_votes = df['Total_Valid_Votes_Clean'].sum()

macro_votes = df[party_columns].sum()
macro_percentages = (macro_votes / total_valid_constituency_votes) * 100

summary_df = pd.DataFrame({
    'Raw_Column': party_columns,
    'Label': [party_config[col]['label'] for col in party_columns],
    'Color': [party_config[col]['color'] for col in party_columns],
    'Total_Votes': macro_votes.values,
    'Share_Percentage': macro_percentages.values
}).sort_values(by='Total_Votes', ascending=True)

fig, ax = plt.subplots(figsize=(14, 8))
macro_bars = ax.barh(summary_df['Label'], summary_df['Share_Percentage'], color=summary_df['Color'], edgecolor='black', height=0.75)

for bar, pct, votes in zip(macro_bars, summary_df['Share_Percentage'], summary_df['Total_Votes']):
    width = bar.get_width()
    ax.text(width + 0.5, bar.get_y() + bar.get_height()/2, f"{int(votes):,} Votes ({pct:.2f}%)", 
            va='center', ha='left', fontsize=8.5, fontweight='bold')

ax.set_title(f"CONSTITUENCY MACRO SUMMARY DASHBOARD\nTotal Valid Votes Cast: {int(total_valid_constituency_votes):,}", fontsize=13, fontweight='bold', pad=25)
ax.set_xlabel('Overall Vote Share Percentage (%)', fontsize=11)
ax.set_xlim(0, max(summary_df['Share_Percentage']) + 15)
ax.invert_yaxis()
plt.tight_layout()

macro_chart_path = f"{output_dir}/Constituency_Macro_Summary.png"
plt.savefig(macro_chart_path, dpi=130, bbox_inches='tight')
plt.close()

# ==============================================================================
# STEP 5: INDIVIDUAL BOOTH VISUALS (SAFE WITH SUFFIX STRINGS)
# ==============================================================================
print("Generating dynamically sorted individual booth graphics...")

for index, row in df.iterrows():
    # Pull directly from matrix index to keep suffixes (like 289A) matching perfectly
    station_id = str(df.at[index, station_col]).strip()
    locality_name = str(row[locality_col]).strip()
    area_name = str(row[area_col]).strip()
    total_votes = row['Total_Valid_Votes_Clean']
    
    if total_votes == 0:
        continue
        
    station_data = pd.DataFrame({
        'Label': [party_config[col]['label'] for col in party_columns],
        'Color': [party_config[col]['color'] for col in party_columns],
        'Votes': row[party_columns].values.astype(float)
    })
    station_data['Percentage'] = (station_data['Votes'] / total_votes) * 100
    station_data = station_data.sort_values(by='Votes', ascending=True)
    
    fig, ax = plt.subplots(figsize=(11, 8.0))
    bars = ax.barh(station_data['Label'], station_data['Percentage'], color=station_data['Color'], edgecolor='black', height=0.65)
    
    for bar, pct, count in zip(bars, station_data['Percentage'], station_data['Votes']):
        width = bar.get_width()
        ax.text(width + 1.2, bar.get_y() + bar.get_height()/2, f"{int(count)} Votes ({pct:.1f}%)", 
                va='center', ha='left', fontsize=8, fontweight='bold')
                
    raw_title_text = f"Station {station_id} | Locality: {locality_name}\nArea: {area_name}"
    wrapped_title = "\n".join(textwrap.wrap(raw_title_text, width=75))
    ax.set_title(f"{wrapped_title} (Total Valid: {int(total_votes)})", fontsize=10, fontweight='bold', pad=15)
    ax.set_xlabel('Vote Share Percentage (%)', fontsize=9)
    ax.set_xlim(0, 140)  
    ax.invert_yaxis()
    
    plt.tight_layout()
    
    safe_locality = "".join([c for c in locality_name if c.isalnum() or c in (' ', '_', '-')]).strip().replace(' ', '_')
    safe_area = "".join([c for c in area_name if c.isalnum() or c in (' ', '_', '-')]).strip().replace(' ', '_')
    
    # Slice the area name string to protect against OS filename length limits (OSError 63)
    short_area = safe_area[:25]
    
    # Filenames format cleanly with your real suffixes (e.g., Station_289A_Locality.png)
    filename = f"{output_dir}/Station_{station_id}_{safe_locality}_{short_area}.png"
    
    plt.savefig(filename, dpi=90, bbox_inches='tight')
    plt.close()

# ==============================================================================
# STEP 6: EXPORT WORKBOOK DATA SHEETS & DYNAMIC CELLS AUTO-FIT
# ==============================================================================
print("Writing analytics tabs and adjusting spreadsheet borders...")
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Detailed_Polling_Data', index=False)
    summary_df[['Label', 'Total_Votes', 'Share_Percentage']].sort_values(by='Total_Votes', ascending=False).to_excel(writer, sheet_name='Constituency_Dashboard', index=False)
    critical_swings.to_excel(writer, sheet_name='Critical_Swing_Booths', index=False)
    locality_dominance_summary.to_excel(writer, sheet_name='Locality_Dominance', index=False)
    independent_vote_splitting.to_excel(writer, sheet_name='Independent_Splitting', index=False)

wb = load_workbook(excel_filename)

ws_dash = wb['Constituency_Dashboard']
img_macro = Image(macro_chart_path)
ws_dash.add_image(img_macro, 'E2')

# FIX: Extracts from row 0 inside the column tuple array to auto-widen cells flawlessly
for sheet in wb.worksheets:
    for col_cells_tuple in sheet.columns:
        max_len = 0
        
        # openpyxl tuple resolution fix: grabs the column code properties from the first cell
        actual_column_index = col_cells_tuple[0].column
        col_letter = get_column_letter(actual_column_index)
        
        for cell in col_cells_tuple:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

wb.save(excel_filename)
print(f"\nSUCCESS! Alphanumeric sorting and openpyxl formatting layout fully resolved: '{excel_filename}'")
import pandas as pd
import numpy as np

# ==============================================================================
# INDEPENDENT SPOILER TRACKING MATRIX
# ==============================================================================
print("Analyzing independent vote-cutting impact on TVK performance...")

# 1. Identify booths where TVK lost but came in a close Second Place (Rank == 2)
# Using your exact config mapping short label 'TVK'
tvk_runner_up_booths = df[df['Tamilaga Vettri Kazhagam_Rank'] == 2].copy()

if tvk_runner_up_booths.empty:
    print("No booths found where TVK finished exactly in 2nd place.")
else:
    # 2. Calculate if the Independent pool was larger than the Margin of Victory
    # If Total_Independent_Votes > Margin_Of_Victory, the independents spoiled the win
    tvk_runner_up_booths['Is_Spoiled_By_Independents'] = (
        tvk_runner_up_booths['Total_Independent_Votes'] > tvk_runner_up_booths['Margin_Of_Victory']
    )
    
    # 3. Filter for the definitive list of Spoiled Booths
    tvk_spoiled_booths = tvk_runner_up_booths[tvk_runner_up_booths['Is_Spoiled_By_Independents'] == True][
        ['Serial No. Of Polling Station', 'Locality', 'Polling Area', 'Winner_Party', 'Winner_Votes', 
         'Tamilaga Vettri Kazhagam', 'Margin_Of_Victory', 'Total_Independent_Votes']
    ].sort_values(by='Margin_Of_Victory', ascending=True)
    
    # Save this custom strategic hit-list back into your Excel report file as a new tab
    with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        tvk_spoiled_booths.to_excel(writer, sheet_name='TVK_Independent_Spoilers', index=False)
        
    print(f"Analysis complete! Found {len(tvk_spoiled_booths)} booths where independents cost TVK the top spot.")
    print(f"Results successfully appended to your spreadsheet under the tab: 'TVK_Independent_Spoilers'")
    
    # Print a quick preview of the top 3 worst-hit booths to your console screen
    if not tvk_spoiled_booths.empty:
        print("\n--- Top 3 Worst Vote-Split Booths for TVK ---")
        print(tvk_spoiled_booths.head(3).to_string(index=False))