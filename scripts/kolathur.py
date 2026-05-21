import os
import textwrap
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# ==============================================================================
# STEP 0: LOAD SOURCE DATA
# ==============================================================================
input_file = "All.csv" 

if not os.path.exists(input_file):
    raise FileNotFoundError(f"Missing base data file. Please ensure '{input_file}' is in this directory.")

print(f"Loading election source dataset: '{input_file}'...")
if input_file.endswith('.csv'):
    df = pd.read_csv(input_file)
else:
    df = pd.read_excel(input_file)

# Clean and normalize column names immediately to avoid string mismatch issues
df.columns = df.columns.str.replace(r'\s+', ' ', regex=True).str.strip().str.lower()

# ==============================================================================
# STEP 1: DEFINE SYSTEM CONFIGURATION (Mapped exactly to your schema)
# ==============================================================================
party_config = {
    'all india anna dravida munnetra kazhagam': {'label': 'AIADMK', 'color': '#2E7D32'},
    'dravida munnetra kazhagam': {'label': 'DMK', 'color': '#D32F2F'},
    'naam tamilar katchi': {'label': 'NTK', 'color': '#FBC02D'},
    'tamilaga vettri kazhagam': {'label': 'TVK', 'color': '#4A148C'},
    'tamizhaga vaazhvurimai katchi': {'label': 'TAVAK', 'color': '#00897B'},
    'bahujan samaj party': {'label': 'BSP', 'color': '#0000FF'},
    'samata party': {'label': 'SAP', 'color': '#FF5722'},
    'makkal nalvaazhvous katchi': {'label': 'MNK', 'color': '#E91E63'},
    'veerath thiyagi viswanathadoss thozhilalarkal katchi': {'label': 'VTVTK', 'color': '#795548'},
    'desiya makkal sakthi katchi': {'label': 'DMSK', 'color': '#9C27B0'},
    'republican party of india (athawale)': {'label': 'RPI(A)', 'color': '#3F51B5'},
    'nota': {'label': 'NOTA', 'color': '#64748B'}, 
}

# Dynamically add all 24 independent columns present in your dataset schema
for i in range(24):
    col_name = 'independent' if i == 0 else f'independent.{i}'
    party_config[col_name] = {'label': f'IND{i+1}', 'color': '#BDBDBD'}

party_columns = [col for col in party_config.keys() if col in df.columns]
independent_cols = [col for col in party_columns if 'independent' in col]

output_dir = 'branded_polling_charts'
os.makedirs(output_dir, exist_ok=True)
excel_filename = 'Branded_Polling_Station_Report.xlsx'

# ==============================================================================
# WORKSPACE CLEANUP & INITIALIZATION
# ==============================================================================
print("Clearing cache layers, dropping untitled columns, and resetting DataFrame...")

columns_to_drop = [f'{col}_rank' for col in party_columns] + [
    'winner_party', 'winner_votes', 'runner_up_votes', 
    'margin_of_victory', 'runner_up_party', 'margin_percentage', 'total_independent_votes',
    'unnamed: 0'
]
df = df.drop(columns=[c for c in columns_to_drop if c in df.columns], errors='ignore')

# ==============================================================================
# SUB-STEP: SPLIT LOCATION AND BUILDING NAMES (Fixed key mapping & string slicing)
# ==============================================================================
print("Splitting Polling Station column into distinct Building and Location layers...")

raw_building_col = 'location and name of building in which polling station located'

def split_polling_station(text):
    if pd.isna(text):
        return "", ""
    text = str(text).strip()
    if ',' in text:
        parts = text.rsplit(',', 1)
        return parts[0].strip(), parts[1].strip()
    return text, text

if raw_building_col in df.columns:
    split_data = df[raw_building_col].apply(split_polling_station)
    df['building_clean'] = [b for b, l in split_data]
    df['location_clean'] = [l for b, l in split_data]
    print("Columns split successfully! Created 'location_clean' and 'building_clean'.")
else:
    print(f"Warning: Column '{raw_building_col}' not found. Skipping split.")
    df['building_clean'] = "N/A"
    df['location_clean'] = "N/A"

# ==============================================================================
# TYPE CASTING & VALUES VALIDATION (Fixed key mapping)
# ==============================================================================
station_col = 'serial no. of polling station'
locality_col = 'location_clean'  
area_col = 'polling area' 

if station_col in df.columns:
    df[station_col] = pd.to_numeric(df[station_col], errors='coerce').fillna(0).astype(int)

vote_cols_to_cast = party_columns + ['total of valid votes', 'nota', 'total', 'no. of rejected votes', 'no. of tendered votes']
for col in vote_cols_to_cast:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
    else:
        print(f"Warning: Configuration column '{col}' not found in the dataset.")

# ==============================================================================
# STEP 2: PRIMARY VOTING CALCULATIONS (WINNER, MARGIN, AND RANKINGS)
# ==============================================================================
print("Running vote analytics calculations...")
active_party_cols = [col for col in party_columns if col in df.columns]

df['Winner_Party'] = df[active_party_cols].idxmax(axis=1)
df['Winner_Votes'] = df[active_party_cols].max(axis=1)

sorted_votes = np.sort(df[active_party_cols].values, axis=1)
df['Runner_Up_Votes'] = sorted_votes[:, -2] if sorted_votes.shape[1] > 1 else 0
df['Margin_Of_Victory'] = df['Winner_Votes'] - df['Runner_Up_Votes']

sorted_indices = np.argsort(df[active_party_cols].values, axis=1)
df['Runner_Up_Party'] = [active_party_cols[idx[-2]] if len(idx) > 1 else 'None' for idx in sorted_indices]

ranks_df = df[active_party_cols].rank(axis=1, ascending=False, method='min').astype(int)
ranks_df = ranks_df.rename(columns={col: f"{party_config[col]['label']}_Rank" for col in active_party_cols})
df = pd.concat([df, ranks_df], axis=1)

# ==============================================================================
# STEP 3: ADVANCED STRATEGIC METRICS
# ==============================================================================
print("Extracting strategic election metrics at individual polling station levels...")

total_votes_col = 'total of valid votes'

df['Margin_Percentage'] = np.where(
    df[total_votes_col] > 0, 
    (df['Margin_Of_Victory'] / df[total_votes_col]) * 100, 
    0.0
)

critical_swings = df[df['Margin_Percentage'] <= 5.0][
    [station_col, locality_col, area_col, 'Winner_Party', 'Runner_Up_Party', 'Margin_Of_Victory', 'Margin_Percentage']
].sort_values(by=station_col, ascending=True).copy()
critical_swings = critical_swings.rename(columns={station_col: 'Polling_Station_No'})

locality_dominance_summary = df[
    [station_col, locality_col, 'Winner_Party', 'Winner_Votes', 'Margin_Of_Victory']
].sort_values(by=station_col, ascending=True).copy()
locality_dominance_summary = locality_dominance_summary.rename(columns={
    station_col: 'Polling_Station_No',
    'Winner_Party': 'Dominant_Party_In_Booth',
    'Winner_Votes': 'Dominant_Party_Votes'
})

active_independent_cols = [col for col in independent_cols if col in df.columns]
df['Total_Independent_Votes'] = df[active_independent_cols].sum(axis=1) if active_independent_cols else 0

independent_vote_splitting = df[
    [station_col, locality_col, total_votes_col, 'Total_Independent_Votes', 'Winner_Party', 'Margin_Of_Victory']
].copy()

independent_vote_splitting['Independent_Vote_Share_%'] = np.where(
    independent_vote_splitting[total_votes_col] > 0,
    (independent_vote_splitting['Total_Independent_Votes'] / independent_vote_splitting[total_votes_col]) * 100,
    0.0
)

independent_vote_splitting['Is_Spoiler_Risk'] = np.where(
    independent_vote_splitting['Total_Independent_Votes'] > independent_vote_splitting['Margin_Of_Victory'],
    'Yes', 'No'
)
independent_vote_splitting = independent_vote_splitting.sort_values(by='Independent_Vote_Share_%', ascending=False)
independent_vote_splitting = independent_vote_splitting.rename(columns={station_col: 'Polling_Station_No'})

df = df.sort_values(by=station_col, ascending=True).reset_index(drop=True)

# ==============================================================================
# STEP 4: GENERATE MACRO CONSTITUENCY DASHBOARD CHART
# ==============================================================================
print("\nCompiling constituency macro-level summary dashboard...")

total_valid_constituency_votes = df[total_votes_col].sum()
macro_votes = df[active_party_cols].sum()
macro_percentages = np.where(total_valid_constituency_votes > 0, (macro_votes / total_valid_constituency_votes) * 100, 0.0)

summary_df = pd.DataFrame({
    'Raw_Column': active_party_cols,
    'Label': [party_config[col]['label'] for col in active_party_cols],
    'Color': [party_config[col]['color'] for col in active_party_cols],
    'Total_Votes': macro_votes.values,
    'Share_Percentage': macro_percentages
}).sort_values(by='Total_Votes', ascending=True)

fig, ax = plt.subplots(figsize=(13, 8))
macro_bars = ax.barh(summary_df['Label'], summary_df['Share_Percentage'], color=summary_df['Color'], edgecolor='black', height=0.7)

for bar, pct, votes in zip(macro_bars, summary_df['Share_Percentage'], summary_df['Total_Votes']):
    width = bar.get_width()
    ax.text(width + 0.8, bar.get_y() + bar.get_height()/2, f"{int(votes):,} Votes ({pct:.2f}%)", va='center', ha='left', fontsize=9, fontweight='bold')

ax.set_title(f"CONSTITUENCY MACRO SUMMARY DASHBOARD\nTotal Valid Votes Cast: {int(total_valid_constituency_votes):,}", fontsize=13, fontweight='bold', pad=20)
ax.set_xlabel('Overall Vote Share Percentage (%)', fontsize=11)
max_pct = max(summary_df['Share_Percentage']) if not summary_df.empty else 0
ax.set_xlim(0, max_pct + 18 if max_pct > 0 else 100)
ax.invert_yaxis()
plt.tight_layout()

macro_chart_path = f"{output_dir}/Constituency_Macro_Summary.png"
plt.savefig(macro_chart_path, dpi=130, bbox_inches='tight')
plt.close()

# ==============================================================================
# STEP 5: INDIVIDUAL BOOTH VISUALS (Fixed String Key and TextWrap Exceptions)
# ==============================================================================
print("Generating dynamically sorted booth images with corrected column mappings...")

for index, row in df.iterrows():
    station_id = str(int(row[station_col]))
    total_votes = row[total_votes_col]
    
    # Use clean split structures generated in step 1 to prevent dynamic crash loops
    building_name = str(row['building_clean']).strip()
    polling_areas_text = str(row[area_col]).strip() if area_col in df.columns else "N/A"
    
    if pd.isna(total_votes) or total_votes == 0:
        continue
        
    station_data = pd.DataFrame({
        'Label': [party_config[col]['label'] for col in active_party_cols],
        'Color': [party_config[col]['color'] for col in active_party_cols],
        'Votes': row[active_party_cols].values.astype(float)
    })
    station_data['Percentage'] = (station_data['Votes'] / total_votes) * 100
    station_data = station_data.sort_values(by='Votes', ascending=True)
    
    fig, ax = plt.subplots(figsize=(11, 8))
    bars = ax.barh(station_data['Label'], station_data['Percentage'], color=station_data['Color'], edgecolor='black', height=0.6)
    
    for bar, pct, count in zip(bars, station_data['Percentage'], station_data['Votes']):
        width = bar.get_width()
        ax.text(
            width + 0.8, 
            bar.get_y() + bar.get_height()/2, 
            f"{int(count)} Votes ({pct:.1f}%)", 
            va='center', 
            ha='left', 
            fontsize=8, 
            fontweight='bold'
        )
                
    wrapped_building = "\n".join(textwrap.wrap(f"Building: {building_name}", width=95))
    wrapped_areas = "\n".join(textwrap.wrap(f"Polling Areas: {polling_areas_text}", width=95))
    
    full_title_text = f"Station {station_id} | Total Votes: {int(total_votes)}\n{wrapped_building}\n{wrapped_areas}"
    
    ax.set_title(full_title_text, fontsize=9, fontweight='bold', pad=18, loc='left')
    ax.set_xlabel('Vote Share Percentage (%)', fontsize=9)
    ax.set_xlim(0, max(station_data['Percentage']) + 20 if max(station_data['Percentage']) > 0 else 115)  
    ax.invert_yaxis()
    plt.tight_layout()
    
    safe_building_slug = "".join([c for c in building_name if c.isalnum() or c in (' ', '_', '-')]).strip().replace(' ', '_')
    plt.savefig(f"{output_dir}/Station_{station_id}_{safe_building_slug[:25]}.png", dpi=95, bbox_inches='tight')
    plt.close()

print(f"All booth graphics generated successfully inside folder: '{output_dir}/'")

# ==============================================================================
# STEP 6: EXPORT TO CLEAN MULTI-SHEET EXCEL WORKBOOK
# ==============================================================================
print("\nAssembling structured spreadsheet file panels...")
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Detailed_Polling_Data', index=False)
    if not summary_df.empty:
        summary_df[['Label', 'Total_Votes', 'Share_Percentage']].sort_values(by='Total_Votes', ascending=False).to_excel(writer, sheet_name='Constituency_Dashboard', index=False)
    if not critical_swings.empty:
        critical_swings.to_excel(writer, sheet_name='Critical_Swing_Booths', index=False)
    if not locality_dominance_summary.empty:
        locality_dominance_summary.to_excel(writer, sheet_name='Locality_Dominance', index=False)
    if not independent_vote_splitting.empty:
        independent_vote_splitting.to_excel(writer, sheet_name='Independent_Splitting', index=False)

wb = load_workbook(excel_filename)
if 'Constituency_Dashboard' in wb.sheetnames and os.path.exists(macro_chart_path):
    ws_dash = wb['Constituency_Dashboard']
    ws_dash.add_image(Image(macro_chart_path), 'E2')

# ==============================================================================
# STEP 7: AUTO-FIT COLUMN WIDTHS ACROSS ALL WORKPASS TABS (Fixed openpyxl Iterator)
# ==============================================================================
print("Optimizing Excel grid padding layout widths...")
for sheet in wb.worksheets:
    for col in list(sheet.columns):
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
                
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

wb.save(excel_filename)
print(f"\nSUCCESS! Branded workbook reporting environment finalized: '{excel_filename}'")
# ==============================================================================
# SPECIALIZED DIAGNOSTIC: WHY DID DMK LOSE IN THIS CONSTITUENCY?
# ==============================================================================
print("\n--- Running DMK Loss Forensic Analysis ---")

dmk_col = 'dravida munnetra kazhagam'
tvk_col = 'tamilaga vettri kazhagam'
aiadmk_col = 'all india anna dravida munnetra kazhagam'

# 1. Overall Constituency Vote Share
total_valid_votes = df[total_votes_col].sum()
dmk_total = df[dmk_col].sum() if dmk_col in df.columns else 0
tvk_total = df[tvk_col].sum() if tvk_col in df.columns else 0
aiadmk_total = df[aiadmk_col].sum() if aiadmk_col in df.columns else 0

dmk_share = (dmk_total / total_valid_votes) * 100 if total_valid_votes > 0 else 0
tvk_share = (tvk_total / total_valid_votes) * 100 if total_valid_votes > 0 else 0
aiadmk_share = (aiadmk_total / total_valid_votes) * 100 if total_valid_votes > 0 else 0

print(f"Constituency Summary:")
print(f"  - DMK Total Vote Share: {dmk_share:.2f}% ({dmk_total:,} votes)")
print(f"  - TVK Total Vote Share: {tvk_share:.2f}% ({tvk_total:,} votes)")
print(f"  - AIADMK Total Vote Share: {aiadmk_share:.2f}% ({aiadmk_total:,} votes)")

# 2. Filter booths where DMK lost
df_dmk_lost = df[df['Winner_Party'] != dmk_col].copy()
total_booths_lost = len(df_dmk_lost)
total_booths = len(df)

print(f"\nDMK lost in {total_booths_lost} out of {total_booths} polling stations.")

if total_booths_lost > 0:
    # 3. Who defeated DMK most often?
    defeating_parties = df_dmk_lost['Winner_Party'].value_counts()
    print("\nParties that won booths away from DMK:")
    for party, count in defeating_parties.items():
        label = party_config.get(party, {}).get('label', party)
        print(f"  - {label}: {count} booths")

    # 4. Independent/Minor Party Vote Splitting Check
    minor_party_cols = [col for col in active_party_cols if col not in [dmk_col, tvk_col, aiadmk_col]]
    df_dmk_lost['Minor_And_Ind_Votes'] = df_dmk_lost[minor_party_cols].sum(axis=1) + df_dmk_lost['Total_Independent_Votes']
    
    # Calculate how far behind DMK was from the winner in each lost booth
    df_dmk_lost['DMK_Deficit'] = df_dmk_lost['Winner_Votes'] - df_dmk_lost[dmk_col]
    
    spoiler_booths = df_dmk_lost[df_dmk_lost['Minor_And_Ind_Votes'] > df_dmk_lost['DMK_Deficit']]
    print(f"\nVote Splitting Impact:")
    print(f"  - In {len(spoiler_booths)} booths, the DMK lost by a margin smaller than the votes wasted on Minor Parties & Independents.")

    # 5. Export this diagnostic to a dedicated sheet including Polling Area
    diagnostic_filename = "DMK_Loss_Analysis.xlsx"
    export_cols = [
        station_col,      # Polling Station Number
        locality_col,     # Clean Building Location
        area_col,         # Polling Area (Added as requested)
        dmk_col,          # DMK Votes
        'Winner_Party',   # Winning Party
        'Winner_Votes',   # Winner's Votes
        'DMK_Deficit',    # Votes needed to win
        'Minor_And_Ind_Votes' # Total spoiler votes
    ]
    
    with pd.ExcelWriter(diagnostic_filename, engine='openpyxl') as diagnostic_writer:
        df_dmk_lost[export_cols].to_excel(diagnostic_writer, sheet_name='Booths_Lost_By_DMK', index=False)
    print(f"\nDetailed booth-by-booth DMK loss breakdown saved to: '{diagnostic_filename}'")
else:
    print("\nDMK won every single booth in this specific dataset. No losses to analyze.")
