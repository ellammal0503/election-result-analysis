import os
import textwrap
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

# ==============================================================================
# STEP 1: DEFINE SYSTEM CONFIGURATION (EXACT MATCH FOR SPECIFIED CANDIDATES)
# ==============================================================================
# Customized brand colors mapping exactly to your active layout specifications
party_config = {
    'Bahujan Samaj Party': {'label': 'BSP', 'color': '#0D47A1'},
    'All India Anna Dravida Munnetra Kazhagam': {'label': 'AIADMK', 'color': '#2E7D32'},
    'Naam Tamilar Katchi': {'label': 'NTK', 'color': '#FBC02D'},
    'Dravida Munnetra Kazhagam': {'label': 'DMK', 'color': '#D32F2F'},
    'Tamilaga Vettri Kazhagam': {'label': 'TVK', 'color': '#7B1E1E'},
    'Indhu Dravida Makkal Katchi': {'label': 'IDMK', 'color': '#E65100'},
    'Our Mahathma National Party': {'label': 'OMNP', 'color': '#6A1B9A'},
    'Tamizhaga Vaazhvurimai Katchi': {'label': 'TAVAK', 'color': '#00897B'},
    'All India Puratchi Thalaivar Makkal Munnettra Kazhagam': {'label': 'AIPTMK', 'color': '#00838F'},
    'Independent': {'label': 'IND', 'color': '#757575'}
}

output_dir = 'branded_polling_charts'
os.makedirs(output_dir, exist_ok=True)
excel_filename = 'Branded_Polling_Station_Report.xlsx'

# ==============================================================================
# WORKSPACE CLEANUP & DATAFRAME INITIALIZATION
# ==============================================================================
print("Clearing cache layers, standardizing indices, and resetting DataFrame...")

# Load your main data sheet here
df = pd.read_csv("All.csv")

# Standardize text layout variables to handle multiple internal spaces safely
df.columns = df.columns.str.replace(r'\s+', ' ', regex=True).str.strip()

# Dynamic Matching Check: Ensure we only isolate parties that exist inside this CSV
master_party_list = list(party_config.keys())
base_party_columns = [col for col in master_party_list if col in df.columns and col != 'Independent']

print(f"🎯 Dynamic Matcher: Found {len(base_party_columns)} explicit party columns in the source dataset.")

# Soft locator fallback strings to handle minor spelling variations across files natively
station_col = [c for c in df.columns if 'station' in c.lower() or 'sl. no' in c.lower()][0]
locality_col = [c for c in df.columns if 'building' in c.lower() or 'location and name' in c.lower()][0]
area_col = [c for c in df.columns if 'area' in c.lower()][0]
voter_type_col = [c for c in df.columns if 'whether' in c.lower()][0]

# Clean up rank, winner, and historical tracking metrics to avoid conflicts
columns_to_drop = [f'{col}_Rank' for col in master_party_list] + [
    'Winner_Party', 'Winner_Votes', 'Runner_Up_Votes', 
    'Margin_Of_Victory', 'Runner_Up_Party', 'Margin_Percentage', 'Total_Other_Candidate_Votes',
    'Unnamed: 0', 'Total_Independent_Votes', 'Calculated_Winner', 'Calculated_Runner_Up', 'Cluster_ID'
]
df = df.drop(columns=[c for c in columns_to_drop if c in df.columns], errors='ignore')

# Convert Serial Numbers to clean integers to eliminate object type mismatch crashes
df[station_col] = pd.to_numeric(df[station_col], errors='coerce').fillna(0).astype('int64')

# Isolate all independent column tracking fields dynamically
non_independent_meta_cols = base_party_columns + [
    station_col, locality_col, area_col, voter_type_col, 'Total of Valid Votes', 
    'No. Of Rejected Votes', 'NOTA', 'Total', 'No. Of Tendered Votes', 'Sl. No'
]
independent_cols = [c for c in df.columns if 'independent' in c.lower() and c not in non_independent_meta_cols]

# Create a clean single Independent aggregator column
if independent_cols:
    df['Independent'] = df[independent_cols].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1)
else:
    if 'Independent' not in df.columns:
        df['Independent'] = 0

# Final verified list of competing structural columns
all_competing_columns = base_party_columns + ['Independent']
df[all_competing_columns] = df[all_competing_columns].apply(pd.to_numeric, errors='coerce').fillna(0)

# ==============================================================================
# STEP 2: PRIMARY VOTING CALCULATIONS (WINNER, MARGIN, AND RANKINGS)
# ==============================================================================
print("Running vote analytics calculations...")

# Pull clean text string array identities directly to avoid unhashable Index crashes
df['Winner_Party'] = df[all_competing_columns].idxmax(axis=1)
df['Winner_Votes'] = df[all_competing_columns].max(axis=1)

# Compute Runner Up metrics cleanly via matrix sorting
sorted_votes = np.sort(df[all_competing_columns].values, axis=1)
df['Runner_Up_Votes'] = sorted_votes[:, -2] if sorted_votes.shape[1] > 1 else sorted_votes[:, -1]
df['Margin_Of_Victory'] = df['Winner_Votes'] - df['Runner_Up_Votes']

# Safely extract Runner Up party labels using positional matrix mapping arrays
sorted_indices = np.argsort(df[all_competing_columns].values, axis=1)
df['Runner_Up_Party'] = [all_competing_columns[idx[-2]] if len(idx) > 1 else all_competing_columns[idx[-1]] for idx in sorted_indices]

# Run localized performance ranks across all competing entries
ranks_df = df[all_competing_columns].rank(axis=1, ascending=False, method='min').astype(int)
ranks_df = ranks_df.rename(columns={col: f'{col}_Rank' for col in all_competing_columns})
df = pd.concat([df, ranks_df], axis=1)

# ==============================================================================
# STEP 3: ADVANCED STRATEGIC METRICS & MACHINE LEARNING K-MEANS ENGINE
# ==============================================================================
print("Extracting strategic election metrics & executing K-Means model...")

df['Margin_Percentage'] = np.where(
    df['Total of Valid Votes'] > 0, 
    (df['Margin_Of_Victory'] / df['Total of Valid Votes']) * 100, 
    0
)

# Identify tightly contested booths (Margin <= 5%)
critical_swings = df[df['Margin_Percentage'] <= 5.0][
    [station_col, locality_col, area_col, 'Winner_Party', 'Runner_Up_Party', 'Margin_Of_Victory', 'Margin_Percentage']
].sort_values(by=station_col, ascending=True)

# Generate a structural analysis matrix of party dominance across localities
locality_counts = df.groupby([locality_col, 'Winner_Party']).size().unstack(fill_value=0)
locality_dominance_summary = pd.DataFrame({
    'Total_Booths_In_Locality': df.groupby(locality_col).size(),
    'Dominant_Party': locality_counts.idxmax(axis=1) if not locality_counts.empty else "None",
    'Dominant_Party_Booths_Won': locality_counts.max(axis=1) if not locality_counts.empty else 0
}).reset_index()

# Track alternative/independent candidate split impacts safely
df['Total_Independent_Votes'] = df['Independent']
independent_vote_splitting = df.groupby(locality_col).agg(
    Total_Valid_Votes_In_Locality=('Total of Valid Votes', 'sum'),
    Total_Independent_Votes_In_Locality=('Total_Independent_Votes', 'sum')
).reset_index()

independent_vote_splitting['Independent_Vote_Share_%'] = np.where(
    independent_vote_splitting['Total_Valid_Votes_In_Locality'] > 0,
    (independent_vote_splitting['Total_Independent_Votes_In_Locality'] / independent_vote_splitting['Total_Valid_Votes_In_Locality']) * 100,
    0
)
independent_vote_splitting = independent_vote_splitting.sort_values(by='Independent_Vote_Share_%', ascending=False)

# K-MEANS CLUSTERING BLOCK
# Generate dynamic share percentages for our active list to use as clustering inputs
cluster_feature_cols = []
for col in all_competing_columns:
    col_name = f"{party_config[col]['label']}_share_pct"
    df[col_name] = np.where(df['Total of Valid Votes'] > 0, (df[col] / df['Total of Valid Votes']) * 100, 0)
    cluster_feature_cols.append(col_name)

cluster_feature_cols.append('Margin_Percentage')
X = df[cluster_feature_cols].copy().fillna(0)

# Scale features and segment booths into 4 optimal strategic cohorts
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X)

kmeans = KMeans(n_clusters=4, init='k-means++', random_state=42, n_init=10)
df['Cluster_ID'] = kmeans.fit_predict(X_scaled)

# Ensure sorting order holds before final visualization steps
df = df.sort_values(by=station_col, ascending=True).reset_index(drop=True)


import os
import textwrap
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import os
import textwrap
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

# ==============================================================================
# STEP 4: GENERATE MACRO CONSTITUENCY DASHBOARD CHART
# ==============================================================================
print("Compiling constituency macro-level summary dashboard...")

total_valid_constituency_votes = df['Total of Valid Votes'].sum()
macro_votes = df[all_competing_columns].sum()

if total_valid_constituency_votes > 0:
    macro_percentages = (macro_votes / total_valid_constituency_votes) * 100
else:
    macro_percentages = macro_votes * 0.0

summary_df = pd.DataFrame({
    'Raw_Column': all_competing_columns,
    'Label': [party_config[col]['label'] for col in all_competing_columns],
    'Color': [party_config[col]['color'] for col in all_competing_columns],
    'Total_Votes': macro_votes.values,
    'Share_Percentage': macro_percentages.values
}).sort_values(by='Total_Votes', ascending=True)

fig, ax = plt.subplots(figsize=(14, 8))
macro_bars = ax.barh(summary_df['Label'], summary_df['Share_Percentage'], color=summary_df['Color'], edgecolor='black', height=0.7)
ax.invert_yaxis()

for bar, pct, votes in zip(macro_bars, summary_df['Share_Percentage'], summary_df['Total_Votes']):
    width = bar.get_width()
    ax.text(width + (summary_df['Share_Percentage'].max() * 0.015), bar.get_y() + bar.get_height()/2, f"{int(votes):,} Votes ({pct:.2f}%)", 
            va='center', ha='left', fontsize=9, fontweight='bold')

ax.set_title(f"CONSTITUENCY MACRO SUMMARY DASHBOARD\nTotal Valid Votes Cast: {int(total_valid_constituency_votes):,}", fontsize=13, fontweight='bold', pad=20)
ax.set_xlabel('Overall Vote Share Percentage (%)', fontsize=11)
ax.set_xlim(0, max(100, summary_df['Share_Percentage'].max() * 1.25))
plt.tight_layout()

macro_chart_path = f"{output_dir}/Constituency_Macro_Summary.png"
plt.savefig(macro_chart_path, dpi=130, bbox_inches='tight')
plt.close()

# ==============================================================================
# STEP 5: INDIVIDUAL BOOTH VISUALS (CLEAN LABELS & COMPACT STORAGE)
# ==============================================================================
print("Generating dynamically sorted booth images...")

for index, row in df.iterrows():
    station_id = str(int(row[station_col]))
    locality_name = str(row[locality_col]).strip()
    area_name = str(row[area_col]).strip()
    total_votes = row['Total of Valid Votes']
    
    if pd.isna(total_votes) or total_votes == 0:
        continue
        
    station_data = pd.DataFrame({
        'Label': [party_config[col]['label'] for col in all_competing_columns],
        'Color': [party_config[col]['color'] for col in all_competing_columns],
        'Votes': row[all_competing_columns].values.astype(float)
    })
    station_data['Percentage'] = (station_data['Votes'] / total_votes) * 100
    station_data = station_data.sort_values(by='Votes', ascending=True)
    
    fig, ax = plt.subplots(figsize=(11.5, 7.0))
    bars = ax.barh(station_data['Label'], station_data['Percentage'], color=station_data['Color'], edgecolor='black', height=0.6)
    ax.invert_yaxis()
    
    for bar, pct, count in zip(bars, station_data['Percentage'], station_data['Votes']):
        width = bar.get_width()
        label_text = f"{int(count)} Votes ({pct:.1f}%)"
        ax.text(width + (station_data['Percentage'].max() * 0.015), bar.get_y() + bar.get_height()/2, label_text, 
                va='center', ha='left', fontsize=8, fontweight='bold')
                
    raw_title_text = f"Station {station_id} | Locality: {locality_name}\nArea: {area_name}"
    wrapped_title = "\n".join(textwrap.wrap(raw_title_text, width=75))
    ax.set_title(f"{wrapped_title} (Total: {int(total_votes)})", fontsize=10, fontweight='bold', pad=15)
    ax.set_xlabel('Vote Share Percentage (%)', fontsize=9)
    ax.set_xlim(0, min(100, station_data['Percentage'].max() * 1.35))  
    
    plt.tight_layout()
    
    safe_locality = "".join([c for c in locality_name if c.isalnum() or c in (' ', '_', '-')]).strip().replace(' ', '_')
    filename = f"{output_dir}/Station_{station_id}_{safe_locality}.png"
    
    plt.savefig(filename, dpi=90, bbox_inches='tight')
    plt.close()

# ==============================================================================
# STEP 6: EXPORT TO CLEAN MULTI-SHEET EXCEL WORKBOOK
# ==============================================================================
print("Assembling structured spreadsheet file panels (No Image Columns)...")
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Detailed_Polling_Data', index=False)
    summary_df[['Label', 'Total_Votes', 'Share_Percentage']].sort_values(by='Total_Votes', ascending=False).to_excel(writer, sheet_name='Constituency_Dashboard', index=False)
    critical_swings.to_excel(writer, sheet_name='Critical_Swing_Booths', index=False)
    locality_dominance_summary.to_excel(writer, sheet_name='Locality_Dominance', index=False)
    independent_vote_splitting.to_excel(writer, sheet_name='Independent_Splitting', index=False)

# Re-load workbook to inject assets via openpyxl drawing tools
wb = load_workbook(excel_filename)

if 'Constituency_Dashboard' in wb.sheetnames:
    ws_dash = wb['Constituency_Dashboard']
    if os.path.exists(macro_chart_path):
        img_macro = Image(macro_chart_path)
        ws_dash.add_image(img_macro, 'E2')

# ==============================================================================
# STEP 7: AUTO-FIT COLUMN WIDTHS & STYLING (FIXED FOR OPENPYXL TUPLES)
# ==============================================================================
print("Optimizing Excel grid padding layout widths and styles...")

for sheet in wb.worksheets:
    # 1. Format headers safely by targeting cell objects individually on Row 1
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # 2. FIXED TUPLE RESOLUTION: Safely inspect columns using cell index boundaries
    for col in sheet.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        
        # FIXED: Extract property attributes safely from the first cell instance inside the tuple array
        actual_column_index = col[0].column
        col_letter = get_column_letter(actual_column_index)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

wb.save(excel_filename)
print(f"\nSUCCESS! Branded reporting workbook environment finalized: '{excel_filename}'")

