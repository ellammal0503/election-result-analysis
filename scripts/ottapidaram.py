import os
import textwrap
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# ==============================================================================
# STEP 1: DEFINE CONFIGURATIONS AND DIRECTORIES
# ==============================================================================
party_config = {
    'Naam Tamilar Katchi': {'label': 'NTK', 'color': '#FBC02D'},
    'Dravida Munnetra Kazhagam': {'label': 'DMK', 'color': '#D32F2F'},
    'Nam Naadu Nam Makkal Nam Ethirkaalam Katchi': {'label': 'NNNMNEK', 'color': '#7B1FA2'},
    'Puthiya Tamilagam': {'label': 'PT', 'color': '#455A64'},
    'Amma Makkal Munnettra Kazagam': {'label': 'AMMK', 'color': '#2E7D32'},
    'Naam Indiar Party': {'label': 'NIP', 'color': '#9E9E9E'},
    'Tamizhaga Vaazhvurimai Katchi': {'label': 'TAVAK', 'color': '#00897B'},
    'Tamilaga Vettri Kazhagam': {'label': 'TVK', 'color': '#7B1E1E'},
    'Independent': {'label': 'IND1', 'color': '#BDBDBD'},
    'Independent.1': {'label': 'IND2', 'color': '#9E9E9E'},
    'Independent.2': {'label': 'IND3', 'color': '#757575'},
    'Independent.3': {'label': 'IND4', 'color': '#616161'},
    'Independent.4': {'label': 'IND5', 'color': '#424242'}
}

master_columns = list(party_config.keys())
output_dir = 'polling_station_charts'
os.makedirs(output_dir, exist_ok=True)

excel_filename = 'Polling_Station_Detailed_Report.xlsx'

# ==============================================================================
# STEP 2: CONSTITUENCY SUMMARY DASHBOARD
# ==============================================================================
print("Generating constituency-wide macro dashboard...")
total_valid_constituency_votes = df['Total of Valid Votes'].sum()

# Aggregate raw total votes for each party across all polling stations
macro_votes = df[master_columns].sum()
macro_percentages = (macro_votes / total_valid_constituency_votes) * 100

# Create a clean lookup dataframe sorted descending for the summary chart
summary_df = pd.DataFrame({
    'Raw_Column': master_columns,
    'Label': [party_config[col]['label'] for col in master_columns],
    'Color': [party_config[col]['color'] for col in master_columns],
    'Total_Votes': macro_votes.values,
    'Share_Percentage': macro_percentages.values
}).sort_values(by='Total_Votes', ascending=True) # Ascending=True plots highest at top in barh

# Plot the macro summary dashboard
fig, ax = plt.subplots(figsize=(12, 6))
macro_bars = ax.barh(summary_df['Label'], summary_df['Share_Percentage'], color=summary_df['Color'], edgecolor='black', height=0.7)

for bar, pct, votes in zip(macro_bars, summary_df['Share_Percentage'], summary_df['Total_Votes']):
    width = bar.get_width()
    ax.text(width + 0.8, bar.get_y() + bar.get_height()/2, f"{int(votes):,} Votes ({pct:.2f}%)", 
            va='center', ha='left', fontsize=10, fontweight='bold')

ax.set_title(f"CONSTITUENCY MACRO SUMMARY DASHBOARD\nTotal Valid Votes Cast: {int(total_valid_constituency_votes):,}", fontsize=14, fontweight='bold', pad=20)
ax.set_xlabel('Overall Vote Share Percentage (%)', fontsize=11)
ax.set_xlim(0, max(summary_df['Share_Percentage']) + 15)
ax.invert_yaxis() # Put the ultimate constituency winner at the absolute top row
plt.tight_layout()

macro_chart_path = f"{output_dir}/Constituency_Macro_Summary.png"
plt.savefig(macro_chart_path, dpi=150, bbox_inches='tight')
plt.close()

# ==============================================================================
# STEP 3: STATION DYNAMIC SORTING & IMAGE GENERATION LOOP
# ==============================================================================
print("Processing individual stations with dynamic ranking layout...")
station_image_mappings = {} # Keeping a registry to know which image matches which Excel row

for index, row in df.iterrows():
    station_id = str(row['Serial No. Of Polling Station']).strip()
    locality_name = str(row['Locality']).strip()
    area_name = str(row['Polling  Area']).strip()
    total_votes = row['Total of Valid Votes']
    
    if pd.isna(total_votes) or total_votes == 0:
        continue
        
    # Build a localized snapshot dataframe for just this single row
    station_data = pd.DataFrame({
        'Label': [party_config[col]['label'] for col in master_columns],
        'Color': [party_config[col]['color'] for col in master_columns],
        'Votes': row[master_columns].values.astype(float)
    })
    station_data['Percentage'] = (station_data['Votes'] / total_votes) * 100
    
    # DYNAMIC SORT: Sort descending by votes so the local station winner is at the top row
    station_data = station_data.sort_values(by='Votes', ascending=True)
    
    fig, ax = plt.subplots(figsize=(11, 5))
    bars = ax.barh(station_data['Label'], station_data['Percentage'], color=station_data['Color'], edgecolor='black', height=0.65)
    
    for bar, pct, count in zip(bars, station_data['Percentage'], station_data['Votes']):
        width = bar.get_width()
        ax.text(width + 1.2, bar.get_y() + bar.get_height()/2, f"{int(count)} Votes ({pct:.1f}%)", 
                va='center', ha='left', fontsize=9, fontweight='bold')
                
    raw_title_text = f"Station {station_id} | Locality: {locality_name} | Polling Area: {area_name}"
    wrapped_title = "\n".join(textwrap.wrap(raw_title_text, width=70))
    ax.set_title(f"{wrapped_title}\n(Total Valid Votes: {int(total_votes)})", fontsize=11, fontweight='bold', pad=20)
    ax.set_xlabel('Vote Share Percentage (%)', fontsize=10)
    ax.set_xlim(0, 135)  
    ax.invert_yaxis() # Forces the local winner of this specific booth to the top
    plt.tight_layout()
    
    # Save with safe filename patterns
    safe_locality = "".join([c for c in locality_name if c.isalnum() or c in (' ', '_', '-')]).strip().replace(' ', '_')
    filename = f"{output_dir}/Station_{station_id}_{safe_locality}.png"
    plt.savefig(filename, dpi=120, bbox_inches='tight') # Reduced DPI slightly so file sizes stay clean inside Excel
    plt.close()
    
    # Register image matching index
    station_image_mappings[index] = filename
# ==============================================================================
# STEP 4: EXPORT STRUCTURAL SHEETS AND EMBED IMAGES INTO EXCEL
# ==============================================================================
print("Assembling reporting sheets and stitching charts into layout cells...")

# Export primary master data layout first to clear file structures
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Detailed_Polling_Data', index=False)
    summary_df[['Label', 'Total_Votes', 'Share_Percentage']].sort_values(by='Total_Votes', ascending=False).to_excel(writer, sheet_name='Constituency_Dashboard', index=False)

# Re-open workbook via openpyxl context engines to append graphics layers
wb = load_workbook(excel_filename)

# Injection Block A: Embed Macro Summary Graphic into Dashboard tab
ws_dash = wb['Constituency_Dashboard']
img_macro = Image(macro_chart_path)
ws_dash.add_image(img_macro, 'E2') 

# Injection Block B: Stitch individual station charts directly into the main data row matrix
ws_data = wb['Detailed_Polling_Data']

# Create a structural column title spacer on the right flank edge for the graphics
graphic_col_idx = len(df.columns) + 2 

# FIX: Set the column header safely without directly copying the Font object
header_cell = ws_data.cell(row=1, column=graphic_col_idx, value="Performance_Visual_Chart")

# Copy individual font properties instead of the parent StyleProxy object to avoid TypeErrors
source_font = ws_data.cell(row=1, column=1).font
if source_font:
    from openpyxl.styles import Font
    header_cell.font = Font(
        name=source_font.name,
        size=source_font.size,
        bold=source_font.bold,
        italic=source_font.italic,
        color=source_font.color
    )

print("Injecting graphic arrays directly into rows. This may take a moment depending on dataset scale...")
for dataframe_idx, image_filepath in station_image_mappings.items():
    excel_row = dataframe_idx + 2 # Offset accounts for 0-index conversion and Excel header row 1
    
    # Set row height to 250 units so the graphic layout box fits elegantly inside the grid row
    ws_data.row_dimensions[excel_row].height = 250
    
    # Convert file paths to openpyxl drawing objects
    img_station = Image(image_filepath)
    
    # Target precise row injection slot coordinates using openpyxl's internal helper
    from openpyxl.utils import get_column_letter
    cell_coordinate = f"{get_column_letter(graphic_col_idx)}{excel_row}"
    ws_data.add_image(img_station, cell_coordinate)

wb.save(excel_filename)
print(f"\nSUCCESS! Ultimate reporting bundle compiled perfectly. Review your new analytical asset: '{excel_filename}'")
