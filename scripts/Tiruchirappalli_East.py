

import os
import textwrap
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# ==============================================================================
# STEP 1: DEFINE SYSTEM CONFIGURATION (EXACT COLUMNS FROM NEW DATASET)
# ==============================================================================
# Refitted to match the active party profile fields in your new 30-column schema
party_config = {
    'Dravida Munnetra Kazhagam': {'label': 'DMK', 'color': '#D32F2F'},
    'All India Anna Dravida Munnetra Kazhagam': {'label': 'AIADMK', 'color': '#2E7D32'},
    'Tamilaga Vettri Kazhagam': {'label': 'TVK', 'color': '#7B1E1E'},
    'Naam Tamilar Katchi': {'label': 'NTK', 'color': '#FBC02D'},
    'Bahujan Samaj Party': {'label': 'BSP', 'color': '#0D47A1'},
    'Puthiya Tamilagam': {'label': 'PT', 'color': '#FF5722'},
    'National Maha Sabha Party': {'label': 'NMSP', 'color': '#757575'},
    'Thamizhaka Padaippalar Makkal Katchi': {'label': 'TPMK', 'color': '#9C27B0'},
    'All India Puratchi Thalaivar Makkal Munnettra Kazhagam': {'label': 'AIPMMK', 'color': '#607D8B'},
    'Samaniya Makkal Nala Katchi': {'label': 'SMNK', 'color': '#795548'},
    'Tamizhaga Vaazhvurimai Katchi': {'label': 'TAVAK', 'color': '#009688'},
    'All India Jananayaka Makkal Kazhagam': {'label': 'AIJMK', 'color': '#E91E63'},
    'Vidial Valarchi Perani': {'label': 'VVP', 'color': '#FF9800'}
}

party_columns = list(party_config.keys())

output_dir = 'branded_polling_charts'
os.makedirs(output_dir, exist_ok=True)
excel_filename = 'Branded_Polling_Station_Report.xlsx'

# ==============================================================================
# WORKSPACE CLEANUP & INITIALIZATION
# ==============================================================================
print("Clearing cache layers, dropping untitled columns, and resetting DataFrame...")

# Force convert multiple inner spacing blocks down to a unified single blank character 
df.columns = df.columns.str.replace(r'\s+', ' ', regex=True).str.strip()

columns_to_drop = [f'{col}_Rank' for col in party_columns] + [
    'Winner_Party', 'Winner_Votes', 'Runner_Up_Votes', 
    'Margin_Of_Victory', 'Runner_Up_Party', 'Margin_Percentage', 'Total_Other_Candidate_Votes',
    'Unnamed: 0', 'Total_Independent_Votes'
]
df = df.drop(columns=[c for c in columns_to_drop if c in df.columns], errors='ignore')

# Dynamic name parameters matched to your single-spaced standard data labels
station_col = 'Serial No. Of Polling Station'
locality_col = 'Location and name of Building in which Polling Station located'
area_col = 'Polling Area'

# Convert Serial Numbers to clean integers to eliminate trailing decimals (.0)
df[station_col] = pd.to_numeric(df[station_col], errors='coerce').fillna(0).astype(int)

# Group structural columns to isolate Independents (Independent to Independent.6)
non_independent_cols = party_columns + [
    'Serial No.', station_col, locality_col, area_col, 'Total of Valid Votes', 
    'No. Of Rejected Votes', 'NOTA', 'Total', 'No. Of Tendered Votes', 'Polling Station Type'
]
independent_cols = [c for c in df.columns if c not in non_independent_cols]

# ==============================================================================
# STEP 2: PRIMARY VOTING CALCULATIONS (WINNER, MARGIN, AND RANKINGS)
# ==============================================================================
print("Running vote analytics calculations...")

df['Winner_Party'] = df[party_columns].idxmax(axis=1)
df['Winner_Votes'] = df[party_columns].max(axis=1)

sorted_votes = np.sort(df[party_columns].values, axis=1)
df['Runner_Up_Votes'] = sorted_votes[:, -2]
df['Margin_Of_Victory'] = df['Winner_Votes'] - df['Runner_Up_Votes']

sorted_indices = np.argsort(df[party_columns].values, axis=1)
df['Runner_Up_Party'] = [party_columns[idx[-2]] for idx in sorted_indices]

ranks_df = df[party_columns].rank(axis=1, ascending=False, method='min').astype(int)
ranks_df = ranks_df.rename(columns={col: f'{col}_Rank' for col in party_columns})
df = pd.concat([df, ranks_df], axis=1)

# ==============================================================================
# STEP 3: ADVANCED STRATEGIC METRICS (SWINGS, DOMINANCE, AND SPOILERS)
# ==============================================================================
print("Extracting strategic election metrics...")

# Added a safeguard for scenarios where Total of Valid Votes is zero
df['Margin_Percentage'] = np.where(
    df['Total of Valid Votes'] > 0, 
    (df['Margin_Of_Victory'] / df['Total of Valid Votes']) * 100, 
    0
)

critical_swings = df[df['Margin_Percentage'] <= 5.0][
    [station_col, locality_col, area_col, 'Winner_Party', 'Runner_Up_Party', 'Margin_Of_Victory', 'Margin_Percentage']
].sort_values(by=station_col, ascending=True)

locality_counts = df.groupby([locality_col, 'Winner_Party']).size().unstack(fill_value=0)
locality_dominance_summary = pd.DataFrame({
    'Total_Booths_In_Locality': df.groupby(locality_col).size(),
    'Dominant_Party': locality_counts.idxmax(axis=1) if not locality_counts.empty else None,
    'Dominant_Party_Booths_Won': locality_counts.max(axis=1) if not locality_counts.empty else 0
}).reset_index()

# Track independent candidate splitting impacts safely across all 7 independent tracking columns
if independent_cols:
    df['Total_Independent_Votes'] = df[independent_cols].sum(axis=1)
    independent_vote_splitting = df.groupby(locality_col).agg(
        Total_Valid_Votes_In_Locality=('Total of Valid Votes', 'sum'),
        Total_Independent_Votes_In_Locality=('Total_Independent_Votes', 'sum')
    ).reset_index()
    independent_vote_splitting['Independent_Vote_Share_%'] = np.where(
        independent_vote_splitting['Total_Valid_Votes_In_Locality'] > 0,
        (independent_vote_splitting['Total_Independent_Votes_In_Locality'] / independent_vote_splitting['Total_Valid_Votes_In_Locality']) * 100,
        0
    )
    independent_vote_splitting = independent_vote_splitting.sort_values(by='Independent_Vote_Share_%', ascending=False)
else:
    print("No independent candidate columns found. Skipping vote splitting shares calculation step.")

# Forcibly sort the master dataset into perfect numerical order by Serial Number before exporting
df = df.sort_values(by=station_col, ascending=True).reset_index(drop=True)
print("Data tracking frame ready for Excel serialization.")

# ==============================================================================
# STEP 4: GENERATE MACRO CONSTITUENCY DASHBOARD CHART
# ==============================================================================
print("Compiling constituency macro-level summary dashboard...")
total_valid_constituency_votes = df['Total of Valid Votes'].sum()

macro_votes = df[party_columns].sum()
macro_percentages = (macro_votes / total_valid_constituency_votes) * 100

summary_df = pd.DataFrame({
    'Raw_Column': party_columns,
    'Label': [party_config[col]['label'] for col in party_columns],
    'Color': [party_config[col]['color'] for col in party_columns],
    'Total_Votes': macro_votes.values,
    'Share_Percentage': macro_percentages.values
}).sort_values(by='Total_Votes', ascending=True)

fig, ax = plt.subplots(figsize=(14, 8))
macro_bars = ax.barh(summary_df['Label'], summary_df['Share_Percentage'], color=summary_df['Color'], edgecolor='black', height=0.7)

# Reverse axis layout to place the leading candidate/party at the very top
ax.invert_yaxis()

for bar, pct, votes in zip(macro_bars, summary_df['Share_Percentage'], summary_df['Total_Votes']):
    width = bar.get_width()
    ax.text(width + (summary_df['Share_Percentage'].max() * 0.015), bar.get_y() + bar.get_height()/2, f"{int(votes):,} Votes ({pct:.2f}%)", 
            va='center', ha='left', fontsize=9, fontweight='bold')

ax.set_title(f"CONSTITUENCY MACRO SUMMARY DASHBOARD\nTotal Valid Votes Cast: {int(total_valid_constituency_votes):,}", fontsize=13, fontweight='bold', pad=20)
ax.set_xlabel('Overall Vote Share Percentage (%)', fontsize=11)
ax.set_xlim(0, summary_df['Share_Percentage'].max() * 1.25)
plt.tight_layout()

macro_chart_path = f"{output_dir}/Constituency_Macro_Summary.png"
plt.savefig(macro_chart_path, dpi=130, bbox_inches='tight')
plt.close()

# ==============================================================================
# STEP 5: INDIVIDUAL BOOTH VISUALS (CLEAN LABELS & COMPACT STORAGE)
# ==============================================================================
print("Generating dynamically sorted booth images...")

for index, row in df.iterrows():
    station_id = str(int(row[station_col]))
    locality_name = str(row[locality_col]).strip()
    area_name = str(row[area_col]).strip()
    total_votes = row['Total of Valid Votes']
    
    if pd.isna(total_votes) or total_votes == 0:
        continue
        
    station_data = pd.DataFrame({
        'Label': [party_config[col]['label'] for col in party_columns],
        'Color': [party_config[col]['color'] for col in party_columns],
        'Votes': row[party_columns].values.astype(float)
    })
    station_data['Percentage'] = (station_data['Votes'] / total_votes) * 100
    station_data = station_data.sort_values(by='Votes', ascending=True)
    
    # Expanded plot height to 7.5 to fit all 13 parties cleanly without overlapping labels
    fig, ax = plt.subplots(figsize=(11.5, 7.5))
    bars = ax.barh(station_data['Label'], station_data['Percentage'], color=station_data['Color'], edgecolor='black', height=0.6)
    
    # Reverse axis layout to place the winning party at the top of the local booth chart
    ax.invert_yaxis()
    
    for bar, pct, count in zip(bars, station_data['Percentage'], station_data['Votes']):
        width = bar.get_width()
        label_text = f"{int(count)} Votes ({pct:.1f}%)"
        ax.text(width + (station_data['Percentage'].max() * 0.015), bar.get_y() + bar.get_height()/2, label_text, 
                va='center', ha='left', fontsize=8, fontweight='bold')
                
    raw_title_text = f"Station {station_id} | Locality: {locality_name}\nArea: {area_name}"
    wrapped_title = "\n".join(textwrap.wrap(raw_title_text, width=75))
    ax.set_title(f"{wrapped_title} (Total: {int(total_votes)})", fontsize=10, fontweight='bold', pad=15)
    ax.set_xlabel('Vote Share Percentage (%)', fontsize=9)
    ax.set_xlim(0, min(100, station_data['Percentage'].max() * 1.35))  
    
    plt.tight_layout()
    
    safe_locality = "".join([c for c in locality_name if c.isalnum() or c in (' ', '_', '-')]).strip().replace(' ', '_')
    filename = f"{output_dir}/Station_{station_id}_{safe_locality}.png"
    
    plt.savefig(filename, dpi=90, bbox_inches='tight')
    plt.close()

# ==============================================================================
# STEP 6: EXPORT TO CLEAN MULTI-SHEET EXCEL WORKBOOK
# ==============================================================================
print("Assembling structured spreadsheet file panels (No Image Columns)...")
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Detailed_Polling_Data', index=False)
    summary_df[['Label', 'Total_Votes', 'Share_Percentage']].sort_values(by='Total_Votes', ascending=False).to_excel(writer, sheet_name='Constituency_Dashboard', index=False)
    critical_swings.to_excel(writer, sheet_name='Critical_Swing_Booths', index=False)
    locality_dominance_summary.to_excel(writer, sheet_name='Locality_Dominance', index=False)
    
    # Conditional export safeguard for independent candidate analysis metrics
    if 'independent_vote_splitting' in locals() or 'independent_vote_splitting' in globals():
        independent_vote_splitting.to_excel(writer, sheet_name='Independent_Splitting', index=False)

# Re-load the freshly written workbook to safely inject assets via openpyxl drawing tools
wb = load_workbook(excel_filename)

# Safely overlay the macro dashboard summary plot onto the primary overview panel
if 'Constituency_Dashboard' in wb.sheetnames:
    ws_dash = wb['Constituency_Dashboard']
    if os.path.exists(macro_chart_path):
        img_macro = Image(macro_chart_path)
        ws_dash.add_image(img_macro, 'E2')
# ==============================================================================
# STEP 7: AUTO-FIT COLUMN WIDTHS & STYLING (FIXED FOR OPENPYXL)
# ==============================================================================
print("Optimizing Excel grid padding layout widths and styles...")
from openpyxl.styles import Font, PatternFill

for sheet in wb.worksheets:
    # 1. FIX: Correctly isolate and loop through individual cells in Row 1 only
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # 2. Fix layout tracking dimensions to auto-adjust width based on maximum content
    for col in sheet.columns:
        max_len = 0
        actual_column_index = col[0].column # Extract column index property safely from the first cell
        col_letter = get_column_letter(actual_column_index)
        
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

wb.save(excel_filename)
print(f"\nSUCCESS! Branded workbook reporting environment finalized: '{excel_filename}'")

