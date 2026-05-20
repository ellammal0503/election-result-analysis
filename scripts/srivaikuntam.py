import os
import textwrap
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# ==============================================================================
# STEP 1: DEFINE SYSTEM CONFIGURATION (ALL 15 ENTITIES)
# ==============================================================================
party_config = {
    'Naam Tamilar Katchi': {'label': 'NTK', 'color': '#FBC02D'},
    'Indian National Congress': {'label': 'INC', 'color': '#D32F2F'}, 
    'Nam Naadu Nam Makkal Nam Ethirkaalam Katchi': {'label': 'NNNMNEK', 'color': '#7B1FA2'},
    'Puthiya Makkal Tamil Desam Katchi': {'label': 'PMTDK', 'color': '#455A64'},
    'All India Anna Dravida Munnetra Kazhagam': {'label': 'AIADMK', 'color': '#2E7D32'}, 
    'Naam Indiar Party': {'label': 'NIP', 'color': '#9E9E9E'},
    'Tamizhaga Vaazhvurimai Katchi': {'label': 'TAVK', 'color': '#00897B'},
    'Tamilaga Vettri Kazhagam': {'label': 'TVK', 'color': '#7B1E1E'},
    'Vishwa Tamil Kazhagam': {'label': 'VTK', 'color': '#FF8C00'},
    'Puthiya Tamilagam': {'label': 'PT', 'color': '#7B1E1E'},
    'All India Forward Bloc': {'label': 'AIFB', 'color': '#D32F2F'}, 
    'Independent': {'label': 'IND1', 'color': '#BDBDBD'},
    'Independent.1': {'label': 'IND2', 'color': '#9E9E9E'},
    'Independent.2': {'label': 'IND3', 'color': '#757575'},
    'Independent.3': {'label': 'IND4', 'color': '#616161'}
}

party_columns = list(party_config.keys())
independent_cols = ['Independent', 'Independent.1', 'Independent.2', 'Independent.3']

output_dir = 'branded_polling_charts'
os.makedirs(output_dir, exist_ok=True)
excel_filename = 'Branded_Polling_Station_Report.xlsx'

# ==============================================================================
# WORKSPACE CLEANUP & DROP UNTITLED COLUMN
# ==============================================================================
print("Clearing cache layers, dropping untitled column, and resetting DataFrame...")
columns_to_drop = [f'{col}_Rank' for col in party_columns] + [
    'Winner_Party', 'Winner_Votes', 'Runner_Up_Votes', 
    'Margin_Of_Victory', 'Runner_Up_Party', 'Margin_Percentage', 'Total_Independent_Votes',
    'Unnamed: 0'  # FIX: Drops the unlabelled index tracker column
]
df = df.drop(columns=[c for c in columns_to_drop if c in df.columns], errors='ignore')

# Convert Serial Numbers to clean integers to eliminate trailing decimals (.0)
df['Serial No. Of Polling Station'] = pd.to_numeric(df['Serial No. Of Polling Station'], errors='coerce').fillna(0).astype(int)

# ==============================================================================
# STEP 2: PRIMARY VOTING CALCULATIONS (WINNER, MARGIN, AND RANKINGS)
# ==============================================================================
print("Running vote analytics calculations...")

df['Winner_Party'] = df[party_columns].idxmax(axis=1)
df['Winner_Votes'] = df[party_columns].max(axis=1)

sorted_votes = np.sort(df[party_columns].values, axis=1)
df['Runner_Up_Votes'] = sorted_votes[:, -2]
df['Margin_Of_Victory'] = df['Winner_Votes'] - df['Runner_Up_Votes']

sorted_indices = np.argsort(df[party_columns].values, axis=1)
df['Runner_Up_Party'] = [party_columns[idx[-2]] for idx in sorted_indices]

ranks_df = df[party_columns].rank(axis=1, ascending=False, method='min').astype(int)
ranks_df = ranks_df.rename(columns={col: f'{col}_Rank' for col in party_columns})
df = pd.concat([df, ranks_df], axis=1)

# ==============================================================================
# STEP 3: ADVANCED STRATEGIC METRICS (SWINGS, DOMINANCE, AND SPOILERS)
# ==============================================================================
print("Extracting strategic election metrics...")

df['Margin_Percentage'] = (df['Margin_Of_Victory'] / df['Total of Valid Votes']) * 100

critical_swings = df[df['Margin_Percentage'] <= 5.0][
    ['Serial No. Of Polling Station', 'Locality', 'Polling  Area', 'Winner_Party', 'Runner_Up_Party', 'Margin_Of_Victory', 'Margin_Percentage']
].sort_values(by='Serial No. Of Polling Station', ascending=True) # Sorted ascending by Station ID

locality_counts = df.groupby(['Locality', 'Winner_Party']).size().unstack(fill_value=0)
locality_dominance_summary = pd.DataFrame({
    'Total_Booths_In_Locality': df.groupby('Locality').size(),
    'Dominant_Party': locality_counts.idxmax(axis=1),
    'Dominant_Party_Booths_Won': locality_counts.max(axis=1)
}).reset_index()

df['Total_Independent_Votes'] = df[independent_cols].sum(axis=1)
independent_vote_splitting = df.groupby('Locality').agg(
    Total_Valid_Votes_In_Locality=('Total of Valid Votes', 'sum'),
    Total_Independent_Votes_In_Locality=('Total_Independent_Votes', 'sum')
).reset_index()
independent_vote_splitting['Independent_Vote_Share_%'] = (
    independent_vote_splitting['Total_Independent_Votes_In_Locality'] / 
    independent_vote_splitting['Total_Valid_Votes_In_Locality']
) * 100
independent_vote_splitting = independent_vote_splitting.sort_values(by='Independent_Vote_Share_%', ascending=False)

# FIX: Forcibly sort the master dataset into perfect numerical order by Serial Number before exporting
df = df.sort_values(by='Serial No. Of Polling Station', ascending=True).reset_index(drop=True)

# ==============================================================================
# STEP 4: GENERATE MACRO CONSTITUENCY DASHBOARD CHART
# ==============================================================================
print("Compiling constituency macro-level summary dashboard...")
total_valid_constituency_votes = df['Total of Valid Votes'].sum()

macro_votes = df[party_columns].sum()
macro_percentages = (macro_votes / total_valid_constituency_votes) * 100

summary_df = pd.DataFrame({
    'Raw_Column': party_columns,
    'Label': [party_config[col]['label'] for col in party_columns],
    'Color': [party_config[col]['color'] for col in party_columns],
    'Total_Votes': macro_votes.values,
    'Share_Percentage': macro_percentages.values
}).sort_values(by='Total_Votes', ascending=True)

fig, ax = plt.subplots(figsize=(13, 7))
macro_bars = ax.barh(summary_df['Label'], summary_df['Share_Percentage'], color=summary_df['Color'], edgecolor='black', height=0.7)

for bar, pct, votes in zip(macro_bars, summary_df['Share_Percentage'], summary_df['Total_Votes']):
    width = bar.get_width()
    ax.text(width + 0.8, bar.get_y() + bar.get_height()/2, f"{int(votes):,} Votes ({pct:.2f}%)", 
            va='center', ha='left', fontsize=9, fontweight='bold')

ax.set_title(f"CONSTITUENCY MACRO SUMMARY DASHBOARD\nTotal Valid Votes Cast: {int(total_valid_constituency_votes):,}", fontsize=13, fontweight='bold', pad=20)
ax.set_xlabel('Overall Vote Share Percentage (%)', fontsize=11)
ax.set_xlim(0, max(summary_df['Share_Percentage']) + 18)
ax.invert_yaxis()
plt.tight_layout()

macro_chart_path = f"{output_dir}/Constituency_Macro_Summary.png"
plt.savefig(macro_chart_path, dpi=130, bbox_inches='tight')
plt.close()

# ==============================================================================
# STEP 5: INDIVIDUAL BOOTH VISUALS (CLEAN VOTE LABEL FORMATTING)
# ==============================================================================
print("Generating dynamically sorted booth images...")

for index, row in df.iterrows():
    station_id = str(int(row['Serial No. Of Polling Station']))
    locality_name = str(row['Locality']).strip()
    area_name = str(row['Polling  Area']).strip()
    total_votes = row['Total of Valid Votes']
    
    if pd.isna(total_votes) or total_votes == 0:
        continue
        
    station_data = pd.DataFrame({
        'Label': [party_config[col]['label'] for col in party_columns],
        'Color': [party_config[col]['color'] for col in party_columns],
        'Votes': row[party_columns].values.astype(float)
    })
    station_data['Percentage'] = (station_data['Votes'] / total_votes) * 100
    station_data = station_data.sort_values(by='Votes', ascending=True)
    
    fig, ax = plt.subplots(figsize=(10.5, 6.0))
    bars = ax.barh(station_data['Label'], station_data['Percentage'], color=station_data['Color'], edgecolor='black', height=0.6)
    
    for bar, pct, count in zip(bars, station_data['Percentage'], station_data['Votes']):
        width = bar.get_width()
        # FIX: Replaced compressed 'v' label with clean, explicit text
        label_text = f"{int(count)} Votes ({pct:.1f}%)"
        ax.text(width + 1.2, bar.get_y() + bar.get_height()/2, label_text, 
                va='center', ha='left', fontsize=8, fontweight='bold')
                
    raw_title_text = f"Station {station_id} | Locality: {locality_name}\nArea: {area_name}"
    wrapped_title = "\n".join(textwrap.wrap(raw_title_text, width=75))
    ax.set_title(f"{wrapped_title} (Total Valid Votes: {int(total_votes)})", fontsize=10, fontweight='bold', pad=15)
    ax.set_xlabel('Vote Share Percentage (%)', fontsize=9)
    ax.set_xlim(0, 140)  
    ax.invert_yaxis()
    
    plt.tight_layout()
    
    safe_locality = "".join([c for c in locality_name if c.isalnum() or c in (' ', '_', '-')]).strip().replace(' ', '_')
    filename = f"{output_dir}/Station_{station_id}_{safe_locality}.png"
    
    plt.savefig(filename, dpi=90, bbox_inches='tight')
    plt.close()

# ==============================================================================
# STEP 6: EXPORT TO CLEAN MULTI-SHEET EXCEL WORKBOOK
# ==============================================================================
print("Assembling structured spreadsheet file panels (No Image Columns)...")
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Detailed_Polling_Data', index=False)
    summary_df[['Label', 'Total_Votes', 'Share_Percentage']].sort_values(by='Total_Votes', ascending=False).to_excel(writer, sheet_name='Constituency_Dashboard', index=False)
    critical_swings.to_excel(writer, sheet_name='Critical_Swing_Booths', index=False)
    locality_dominance_summary.to_excel(writer, sheet_name='Locality_Dominance', index=False)
    independent_vote_splitting.to_excel(writer, sheet_name='Independent_Splitting', index=False)

wb = load_workbook(excel_filename)

# Embed core dashboard visual block into its dedicated tab
ws_dash = wb['Constituency_Dashboard']
img_macro = Image(macro_chart_path)
ws_dash.add_image(img_macro, 'E2')

# ==============================================================================
# STEP 7: AUTO-FIT COLUMN WIDTHS ACROSS ALL WORKPASS TABS
# ==============================================================================
print("Optimizing Excel grid padding layout widths...")
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        actual_column_index = col[0].column # FIXED: Reads column property directly from the collection cells wrapper
        col_letter = get_column_letter(actual_column_index)
        
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

wb.save(excel_filename)
print(f"\nSUCCESS! Execution pipeline finalized flawlessly. File written to: '{excel_filename}'")
