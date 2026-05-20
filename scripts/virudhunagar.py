import os
import re
import textwrap
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# ==============================================================================
# STEP 1: DEFINE SYSTEM CONFIGURATION (RE-ALIGNED FOR NEW SCHEMA)
# ==============================================================================
party_config = {
    'All India Anna Dravida Munnetra Kazhagam': {'label': 'AIADMK', 'color': '#2E7D32'},
    'Naam Tamilar Katchi': {'label': 'NTK', 'color': '#FBC02D'},
    'Desiya Murpokku Dravida Kazhagam': {'label': 'DMDK', 'color': '#FF9800'},
    'Tamilaga Vettri Kazhagam': {'label': 'TVK', 'color': '#7B1E1E'},
    'Puthiya Tamilagam': {'label': 'PT', 'color': '#00897B'},
    'Nam Naadu Nam Makkal Nam Ethirkaalam Katchi': {'label': 'NNNMNEK', 'color': '#424242'},
    'Desa Makkal Munnetrak Kazhagam': {'label': 'DMMK', 'color': '#455A64'},
    'Independent': {'label': 'IND1', 'color': '#BDBDBD'},
    'Independent.1': {'label': 'IND2', 'color': '#9E9E9E'},
    'Independent.2': {'label': 'IND3', 'color': '#757575'},
    'Independent.3': {'label': 'IND4', 'color': '#616161'},
    'Independent.4': {'label': 'IND5', 'color': '#424242'},
    'Independent.5': {'label': 'IND6', 'color': '#EEEEEE'},
    'Independent.6': {'label': 'IND7', 'color': '#E0E0E0'},
    'Independent.7': {'label': 'IND8', 'color': '#CCCCCC'}
}

party_columns = list(party_config.keys())
independent_cols = [col for col in party_columns if 'Independent' in col]

output_dir = 'branded_polling_charts_new'
os.makedirs(output_dir, exist_ok=True)
excel_filename = 'Branded_Polling_Station_Report_New.xlsx'

# ==============================================================================
# WORKSPACE CLEANUP & INITIALIZATION (NO SPLITTING - NATIVE COLUMNS ONLY)
# ==============================================================================
print("Cleaning text labels and stripping hidden spaces...")
df.columns = df.columns.str.replace(r'\s+', ' ', regex=True).str.strip()

# Clean up redundant split-leftovers if they exist in your source variable layers
columns_to_drop = [f'{col}_Rank' for col in party_columns] + [
    'Winner_Party', 'Winner_Votes', 'Runner_Up_Votes', 
    'Margin_Of_Victory', 'Runner_Up_Party', 'Margin_Percentage', 'Total_Independent_Votes',
    'Building in Which Polling Station Located', 'location', 'Unnamed: 0'
]
df = df.drop(columns=[c for c in columns_to_drop if c in df.columns], errors='ignore')

# Extract exact un-wrapped string names directly out of the matched arrays
station_col = [c for c in df.columns if 'Serial' in str(c) or 'Station' in str(c)][0]
locality_col = [c for c in df.columns if 'Building' in str(c) or 'Location' in str(c)][0]
area_col = [c for c in df.columns if 'Area' in str(c)][0]
type_col = [c for c in df.columns if 'Type' in str(c)][0]

# Convert numeric columns safely
df[station_col] = pd.to_numeric(df[station_col], errors='coerce').fillna(0).astype(int)
for col in party_columns + ['Total of Valid Votes']:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

# ==============================================================================
# STEP 2: PRIMARY VOTING CALCULATIONS
# ==============================================================================
print("Running fundamental voting calculations...")
df['Winner_Party'] = df[party_columns].idxmax(axis=1)
df['Winner_Votes'] = df[party_columns].max(axis=1)

sorted_votes = np.sort(df[party_columns].values, axis=1)
df['Runner_Up_Votes'] = sorted_votes[:, -2]
df['Margin_Of_Victory'] = df['Winner_Votes'] - df['Runner_Up_Votes']

sorted_indices = np.argsort(df[party_columns].values, axis=1)
df['Runner_Up_Party'] = [party_columns[idx[-2]] for idx in sorted_indices]

ranks_df = df[party_columns].rank(axis=1, ascending=False, method='min').astype(int)
ranks_df = ranks_df.rename(columns={col: f'{col}_Rank' for col in party_columns})
df = pd.concat([df, ranks_df], axis=1)

# ==============================================================================
# STEP 3: STRATEGIC METRICS & TVK DIAGNOSTICS (WITH STATION LIST MAPPING)
# ==============================================================================
print("Extracting margin swings, spoiler tracking, and TVK diagnostic panels...")
df['Margin_Percentage'] = np.where(df['Total of Valid Votes'] > 0, (df['Margin_Of_Victory'] / df['Total of Valid Votes']) * 100, 0.0)

critical_swings = df[df['Margin_Percentage'] <= 5.0][
    [station_col, locality_col, area_col, type_col, 'Winner_Party', 'Runner_Up_Party', 'Margin_Of_Victory', 'Margin_Percentage']
].sort_values(by=station_col)

# 1. Locality Dominance Summary (Pivoted to Area with Station Serial Numbers mapped)
area_counts = df.groupby([area_col, 'Winner_Party']).size().unstack(fill_value=0)
locality_dominance_summary = pd.DataFrame({
    'Polling_Stations_List': df.groupby(area_col)[station_col].apply(lambda x: ', '.join(x.astype(str))),
    'Total_Booths_In_Area': df.groupby(area_col).size(),
    'Dominant_Party': area_counts.idxmax(axis=1),
    'Dominant_Party_Booths_Won': area_counts.max(axis=1)
}).reset_index()

df['Total_Independent_Votes'] = df[independent_cols].sum(axis=1)

# 2. Independent Vote Share Splitting Summary (With Station Serial Numbers mapped)
independent_vote_splitting = df.groupby(area_col).agg(
    Polling_Stations_List=(station_col, lambda x: ', '.join(x.astype(str))),
    Total_Valid_Votes_In_Area=('Total of Valid Votes', 'sum'),
    Total_Independent_Votes_In_Area=('Total_Independent_Votes', 'sum')
).reset_index()

independent_vote_splitting['Independent_Vote_Share_%'] = np.where(
    independent_vote_splitting['Total_Valid_Votes_In_Area'] > 0,
    (independent_vote_splitting['Total_Independent_Votes_In_Area'] / independent_vote_splitting['Total_Valid_Votes_In_Area']) * 100,
    0.0
)
independent_vote_splitting = independent_vote_splitting.sort_values(by='Independent_Vote_Share_%', ascending=False)

# --- TVK SPECIFIC CONSTITUENCY ANALYSIS PANEL ---
print("\n" + "="*50 + "\nLOCAL TVK PERFORMANCE POST-MORTEM\n" + "="*50)
tvk_won = df[df['Winner_Party'] == 'Tamilaga Vettri Kazhagam']
spoiler_booths = df[
    (df['Winner_Party'] != 'Tamilaga Vettri Kazhagam') & 
    (df['Tamilaga Vettri Kazhagam_Rank'] == 2) & 
    (df['Total_Independent_Votes'] > df['Margin_Of_Victory'])
]
print(f"Total Polling Stations Analyzed: {len(df)}")
print(f"Booths Won by TVK              : {len(tvk_won)} ({ (len(tvk_won)/len(df))*100 :.2f}%)")
print(f"Booths lost due to IND Spoilers: {len(spoiler_booths)}")
print("="*50 + "\n")

df = df.sort_values(by=station_col).reset_index(drop=True)

# ==============================================================================
# STEP 4: MACRO DASHBOARD PLOT
# ==============================================================================
print("Compiling constituency macro dashboard visual chart...")
total_valid_constituency_votes = df['Total of Valid Votes'].sum()
macro_votes = df[party_columns].sum()
macro_percentages = np.where(total_valid_constituency_votes > 0, (macro_votes / total_valid_constituency_votes) * 100, 0.0)

summary_df = pd.DataFrame({
    'Raw_Column': party_columns,
    'Label': [party_config[col]['label'] for col in party_columns],
    'Color': [party_config[col]['color'] for col in party_columns],
    'Total_Votes': macro_votes.values,
    'Share_Percentage': macro_percentages
}).sort_values(by='Total_Votes', ascending=True)

fig, ax = plt.subplots(figsize=(13, 8))
macro_bars = ax.barh(summary_df['Label'], summary_df['Share_Percentage'], color=summary_df['Color'], edgecolor='black', height=0.7)

for bar, pct, votes in zip(macro_bars, summary_df['Share_Percentage'], summary_df['Total_Votes']):
    width = bar.get_width()
    ax.text(width + 0.8, bar.get_y() + bar.get_height()/2, f"{int(votes):,} ({pct:.2f}%)", va='center', ha='left', fontsize=9, fontweight='bold')

ax.set_title(f"CONSTITUENCY MACRO SUMMARY\nTotal Votes: {int(total_valid_constituency_votes):,}", fontsize=13, fontweight='bold', pad=20)
ax.set_xlim(0, max(summary_df['Share_Percentage']) + 18 if max(summary_df['Share_Percentage']) > 0 else 115)
ax.invert_yaxis()
plt.tight_layout()
macro_chart_path = f"{output_dir}/Macro_Summary.png"
plt.savefig(macro_chart_path, dpi=130, bbox_inches='tight')
plt.close()

# ==============================================================================
# STEP 5: INDIVIDUAL BOOTH PLOTS (CUSTOM UN-SPLIT FORMAT WITH FULL AREA EXPOSURE)
# ==============================================================================
print("Generating booth visual images with custom double-underscore text processing...")

def make_filename_safe(text):
    if pd.isna(text):
        return ""
    s = re.sub(r'\s+', ' ', str(text)).strip()
    s = re.sub(r'[^a-zA-Z0-9]', '_', s)
    s = re.sub(r'_{2,}', '__', s)
    return s.strip('_')

for index, row in df.iterrows():
    station_id = str(int(row[station_col]))
    building_full_name = str(row[locality_col]).strip()
    full_polling_area = str(row[area_col]).strip()
    booth_type = str(row[type_col]).strip()
    total_votes = row['Total of Valid Votes']
    
    if pd.isna(total_votes) or total_votes == 0:
        continue
        
    station_data = pd.DataFrame({
        'Label': [party_config[col]['label'] for col in party_columns],
        'Color': [party_config[col]['color'] for col in party_columns],
        'Votes': row[party_columns].values.astype(float)
    })
    station_data['Percentage'] = (station_data['Votes'] / total_votes) * 100
    station_data = station_data.sort_values(by='Votes', ascending=True)
    
    fig, ax = plt.subplots(figsize=(11, 7.5))
    bars = ax.barh(station_data['Label'], station_data['Percentage'], color=station_data['Color'], edgecolor='black', height=0.6)
    
    for bar, pct, count in zip(bars, station_data['Percentage'], station_data['Votes']):
        width = bar.get_width()
        ax.text(width + 0.8, bar.get_y() + bar.get_height()/2, f"{int(count)} ({pct:.1f}%)", va='center', ha='left', fontsize=8, fontweight='bold')
                
    wrapped_building = "\n".join(textwrap.wrap(f"Station {station_id} [{booth_type}] | {building_full_name}", width=85))
    wrapped_area = "\n".join(textwrap.wrap(f"Polling Areas: {full_polling_area}", width=95))
    
    ax.set_title(f"{wrapped_building}\n{wrapped_area}\n(Total Votes Cast: {int(total_votes)})", fontsize=9, fontweight='bold', pad=15)
    ax.set_xlim(0, max(station_data['Percentage']) + 20 if max(station_data['Percentage']) > 0 else 115)  
    ax.invert_yaxis()
    plt.tight_layout()
    
    safe_building_string = make_filename_safe(building_full_name)
    image_filename = f"{output_dir}/Station_{station_id}_{safe_building_string}__.png"
    
    if len(image_filename) > 242:
        image_filename = f"{output_dir}/Station_{station_id}_{safe_building_string[:180]}__.png"
        
    plt.savefig(image_filename, dpi=90, bbox_inches='tight')
    plt.close()

# ==============================================================================
# STEPS 6 & 7: EXCEL GENERATION & WIDTH AUTO-FIT (FIXED TUPLE RUNTIME)
# ==============================================================================
print("Assembling workbook spreadsheet file panels...")
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Detailed_Polling_Data', index=False)
    summary_df[['Label', 'Total_Votes', 'Share_Percentage']].sort_values(by='Total_Votes', ascending=False).to_excel(writer, sheet_name='Constituency_Dashboard', index=False)
    critical_swings.to_excel(writer, sheet_name='Critical_Swing_Booths', index=False)
    locality_dominance_summary.to_excel(writer, sheet_name='Locality_Dominance', index=False)
    independent_vote_splitting.to_excel(writer, sheet_name='Independent_Splitting', index=False)

wb = load_workbook(excel_filename)
wb['Constituency_Dashboard'].add_image(Image(macro_chart_path), 'E2')

print("Optimizing Excel grid layout padding widths...")
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        actual_column_index = col[0].column 
        col_letter = get_column_letter(actual_column_index)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

wb.save(excel_filename)
print(f"\nSUCCESS! New report processed and saved to: '{excel_filename}'")
