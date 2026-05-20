import os
import textwrap
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# ==============================================================================
# STEP 1: DEFINE SYSTEM CONFIGURATION (MATCHED TO YOUR EXACT DATASET)
# ==============================================================================
party_config = {
    'Dravida Munnetra Kazhagam': {'label': 'DMK', 'color': '#D32F2F'},
    'All India Anna Dravida Munnetra Kazhagam': {'label': 'AIADMK', 'color': '#2E7D32'},
    'Naam Tamilar Katchi': {'label': 'NTK', 'color': '#FBC02D'},
    'Puthiya Makkal Tamil Desam Katchi': {'label': 'PMTDK', 'color': '#FF9800'},
    'Tamizhaga Vaazhvurimai Katchi': {'label': 'TAVAK', 'color': '#00897B'},
    'Naam Indiar Party': {'label': 'NIP', 'color': '#9E9E9E'},
    'Tamilaga Vettri Kazhagam': {'label': 'TVK', 'color': '#7B1E1E'},
    'Independent': {'label': 'IND1', 'color': '#BDBDBD'},
    'Independent.1': {'label': 'IND2', 'color': '#9E9E9E'},
    'Independent.2': {'label': 'IND3', 'color': '#757575'},
    'Independent.3': {'label': 'IND4', 'color': '#616161'},
    'Independent.4': {'label': 'IND5', 'color': '#424242'},
    'Independent.5': {'label': 'IND6', 'color': '#EEEEEE'},
    'Independent.6': {'label': 'IND7', 'color': '#E0E0E0'},
    'Independent.7': {'label': 'IND8', 'color': '#CCCCCC'},
    'Independent.8': {'label': 'IND9', 'color': '#B0B0B0'},
    'Independent.9': {'label': 'IND10', 'color': '#9A9A9A'}
}

party_columns = list(party_config.keys())
independent_cols = [col for col in party_columns if 'Independent' in col]

output_dir = 'branded_polling_charts'
os.makedirs(output_dir, exist_ok=True)
excel_filename = 'Branded_Polling_Station_Report.xlsx'

# ==============================================================================
# WORKSPACE CLEANUP & INITIALIZATION
# ==============================================================================
print("Clearing cache layers, dropping untitled column, and resetting DataFrame...")
df.columns = df.columns.str.replace(r'\s+', ' ', regex=True).str.strip()
# ==============================================================================
# SUB-STEP: SPLIT LOCATION AND BUILDING NAMES
# ==============================================================================
print("Splitting Polling Station column into distinct Building and Location layers...")

# Identify the exact column name dynamically from your schema
raw_building_col = [c for c in df.columns if 'Building' in str(c) or 'Location' in str(c)][0]

def split_polling_station(text):
    if pd.isna(text):
        return "", ""
    
    text = str(text).strip()
    
    # Split from the right side at the very last comma
    if ',' in text:
        parts = text.rsplit(',', 1)
        building = parts[0].strip()
        location = parts[1].strip()
        return building, location
    
    # Fallback if no comma exists
    return text, text

# Apply the split logic to generate two clean, separate columns
split_data = df[raw_building_col].apply(split_polling_station)
df['Building in Which Polling Station Located'] = [b for b, l in split_data]
df['location'] = [l for b, l in split_data]

print("Columns split successfully! Created 'location' and 'Building in Which Polling Station Located'.")

columns_to_drop = [f'{col}_Rank' for col in party_columns] + [
    'Winner_Party', 'Winner_Votes', 'Runner_Up_Votes', 
    'Margin_Of_Victory', 'Runner_Up_Party', 'Margin_Percentage', 'Total_Independent_Votes',
    'Unnamed: 0'
]
df = df.drop(columns=[c for c in columns_to_drop if c in df.columns], errors='ignore')

# Dynamic resolution mapping using index list extraction 
station_col = [c for c in df.columns if 'Serial No.' in str(c) or 'Station' in str(c)][0]
locality_col = [c for c in df.columns if 'Building' in str(c) or 'Located' in str(c)][0]
area_col = [c for c in df.columns if 'Areas' in str(c) or 'Area' in str(c)][0]

df[station_col] = pd.to_numeric(df[station_col], errors='coerce').fillna(0).astype(int)

for col in party_columns + ['Total of Valid Votes']:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

# ==============================================================================
# STEP 2: PRIMARY VOTING CALCULATIONS (WINNER, MARGIN, AND RANKINGS)
# ==============================================================================
print("Running vote analytics calculations...")
df['Winner_Party'] = df[party_columns].idxmax(axis=1)
df['Winner_Votes'] = df[party_columns].max(axis=1)

sorted_votes = np.sort(df[party_columns].values, axis=1)
df['Runner_Up_Votes'] = sorted_votes[:, -2]
df['Margin_Of_Victory'] = df['Winner_Votes'] - df['Runner_Up_Votes']

sorted_indices = np.argsort(df[party_columns].values, axis=1)
df['Runner_Up_Party'] = [party_columns[idx[-2]] for idx in sorted_indices]

ranks_df = df[party_columns].rank(axis=1, ascending=False, method='min').astype(int)
ranks_df = ranks_df.rename(columns={col: f'{col}_Rank' for col in party_columns})
df = pd.concat([df, ranks_df], axis=1)

# ==============================================================================
# STEP 3: ADVANCED STRATEGIC METRICS (SWINGS, DOMINANCE, AND SPOILERS)
# ==============================================================================
print("Extracting strategic election metrics...")
df['Margin_Percentage'] = np.where(df['Total of Valid Votes'] > 0, (df['Margin_Of_Victory'] / df['Total of Valid Votes']) * 100, 0.0)

critical_swings = df[df['Margin_Percentage'] <= 5.0][
    [station_col, locality_col, area_col, 'Winner_Party', 'Runner_Up_Party', 'Margin_Of_Victory', 'Margin_Percentage']
].sort_values(by=station_col, ascending=True)

locality_counts = df.groupby([locality_col, 'Winner_Party']).size().unstack(fill_value=0)
locality_dominance_summary = pd.DataFrame({
    'Total_Booths_In_Locality': df.groupby(locality_col).size(),
    'Dominant_Party': locality_counts.idxmax(axis=1),
    'Dominant_Party_Booths_Won': locality_counts.max(axis=1)
}).reset_index()

df['Total_Independent_Votes'] = df[independent_cols].sum(axis=1)
independent_vote_splitting = df.groupby(locality_col).agg(
    Total_Valid_Votes_In_Locality=('Total of Valid Votes', 'sum'),
    Total_Independent_Votes_In_Locality=('Total_Independent_Votes', 'sum')
).reset_index()

independent_vote_splitting['Independent_Vote_Share_%'] = np.where(
    independent_vote_splitting['Total_Valid_Votes_In_Locality'] > 0,
    (independent_vote_splitting['Total_Independent_Votes_In_Locality'] / independent_vote_splitting['Total_Valid_Votes_In_Locality']) * 100,
    0.0
)
independent_vote_splitting = independent_vote_splitting.sort_values(by='Independent_Vote_Share_%', ascending=False)
df = df.sort_values(by=station_col, ascending=True).reset_index(drop=True)

# ==============================================================================
# TVK SPECIFIC PERFORMANCE DIAGNOSTICS
# ==============================================================================
print("\n=== TVK POST-MORTEM DATA ANALYSIS ===")

# 1. Macro Win/Loss Counts
tvk_won = df[df['Winner_Party'] == 'Tamilaga Vettri Kazhagam']
tvk_lost = df[df['Winner_Party'] != 'Tamilaga Vettri Kazhagam']

print(f"Total Polling Stations Analyzed: {len(df)}")
print(f"Booths Won by TVK: {len(tvk_won)} ({ (len(tvk_won)/len(df))*100 :.2f}%)")
print(f"Booths Lost by TVK: {len(tvk_lost)} ({ (len(tvk_lost)/len(df))*100 :.2f}%)")

# 2. Identify Heartbreak Booths (TVK came 2nd by a razor-thin margin)
tvk_runner_up_tight = df[
    (df['Runner_Up_Party'] == 'Tamilaga Vettri Kazhagam') & 
    (df['Margin_Of_Victory'] <= 15)
]
print(f"Heartbreak Booths (TVK lost by 15 votes or less): {len(tvk_runner_up_tight)}")

# 3. Independent Candidates Spoiler Effect
# Checking booths where TVK lost, but the Independent vote pool was greater than the losing margin
spoiler_booths = df[
    (df['Winner_Party'] != 'Tamilaga Vettri Kazhagam') & 
    (df['Tamilaga Vettri Kazhagam_Rank'] == 2) & 
    (df['Total_Independent_Votes'] > df['Margin_Of_Victory'])
]
print(f"Booths where Independent Candidates cost TVK the win: {len(spoiler_booths)}")

# 4. Identify Toughest Territory (Where TVK collapsed to 3rd place or lower)
tvk_underperformed = df[df['Tamilaga Vettri Kazhagam_Rank'] >= 3]
if not tvk_underperformed.empty:
    worst_locality = tvk_underperformed.groupby(locality_col[0]).size().idxmax()
    print(f"Locality with lowest TVK penetration (Most 3rd+ place finishes): {worst_locality}")


# ==============================================================================
# STEP 4: GENERATE MACRO CONSTITUENCY DASHBOARD CHART
# ==============================================================================
print("Compiling constituency macro-level summary dashboard...")
total_valid_constituency_votes = df['Total of Valid Votes'].sum()
macro_votes = df[party_columns].sum()
macro_percentages = np.where(total_valid_constituency_votes > 0, (macro_votes / total_valid_constituency_votes) * 100, 0.0)

summary_df = pd.DataFrame({
    'Raw_Column': party_columns,
    'Label': [party_config[col]['label'] for col in party_columns],
    'Color': [party_config[col]['color'] for col in party_columns],
    'Total_Votes': macro_votes.values,
    'Share_Percentage': macro_percentages
}).sort_values(by='Total_Votes', ascending=True)

fig, ax = plt.subplots(figsize=(13, 8))
macro_bars = ax.barh(summary_df['Label'], summary_df['Share_Percentage'], color=summary_df['Color'], edgecolor='black', height=0.7)

for bar, pct, votes in zip(macro_bars, summary_df['Share_Percentage'], summary_df['Total_Votes']):
    width = bar.get_width()
    ax.text(width + 0.8, bar.get_y() + bar.get_height()/2, f"{int(votes):,} Votes ({pct:.2f}%)", va='center', ha='left', fontsize=9, fontweight='bold')

ax.set_title(f"CONSTITUENCY MACRO SUMMARY DASHBOARD\nTotal Valid Votes Cast: {int(total_valid_constituency_votes):,}", fontsize=13, fontweight='bold', pad=20)
ax.set_xlabel('Overall Vote Share Percentage (%)', fontsize=11)
ax.set_xlim(0, max(summary_df['Share_Percentage']) + 18 if max(summary_df['Share_Percentage']) > 0 else 115)
ax.invert_yaxis()
plt.tight_layout()

macro_chart_path = f"{output_dir}/Constituency_Macro_Summary.png"
plt.savefig(macro_chart_path, dpi=130, bbox_inches='tight')
plt.close()

# ==============================================================================
# STEP 5: INDIVIDUAL BOOTH VISUALS (CLEAN LABELS & COMPACT STORAGE)
# ==============================================================================
print("Generating dynamically sorted booth images...")
for index, row in df.iterrows():
    station_id = str(int(row[station_col]))
    locality_name = str(row[locality_col]).strip()
    area_name = str(row[area_col]).strip()
    total_votes = row['Total of Valid Votes']
    
    if pd.isna(total_votes) or total_votes == 0:
        continue
        
    station_data = pd.DataFrame({
        'Label': [party_config[col]['label'] for col in party_columns],
        'Color': [party_config[col]['color'] for col in party_columns],
        'Votes': row[party_columns].values.astype(float)
    })
    station_data['Percentage'] = (station_data['Votes'] / total_votes) * 100
    station_data = station_data.sort_values(by='Votes', ascending=True)
    
    fig, ax = plt.subplots(figsize=(11, 7))
    bars = ax.barh(station_data['Label'], station_data['Percentage'], color=station_data['Color'], edgecolor='black', height=0.6)
    
    for bar, pct, count in zip(bars, station_data['Percentage'], station_data['Votes']):
        width = bar.get_width()
        ax.text(width + 0.8, bar.get_y() + bar.get_height()/2, f"{int(count)} Votes ({pct:.1f}%)", va='center', ha='left', fontsize=8, fontweight='bold')
                
    wrapped_title = "\n".join(textwrap.wrap(f"Station {station_id} | Locality: {locality_name}\nArea: {area_name}", width=80))
    ax.set_title(f"{wrapped_title} (Total: {int(total_votes)})", fontsize=10, fontweight='bold', pad=15)
    ax.set_xlabel('Vote Share Percentage (%)', fontsize=9)
    ax.set_xlim(0, max(station_data['Percentage']) + 20 if max(station_data['Percentage']) > 0 else 115)  
    ax.invert_yaxis()
    plt.tight_layout()
    
    safe_locality = "".join([c for c in locality_name if c.isalnum() or c in (' ', '_', '-')]).strip().replace(' ', '_')
    plt.savefig(f"{output_dir}/Station_{station_id}_{safe_locality[:30]}.png", dpi=90, bbox_inches='tight')
    plt.close()

# ==============================================================================
# STEP 6: EXPORT TO CLEAN MULTI-SHEET EXCEL WORKBOOK
# ==============================================================================
print("Assembling structured spreadsheet file panels (No Image Columns)...")
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Detailed_Polling_Data', index=False)
    summary_df[['Label', 'Total_Votes', 'Share_Percentage']].sort_values(by='Total_Votes', ascending=False).to_excel(writer, sheet_name='Constituency_Dashboard', index=False)
    critical_swings.to_excel(writer, sheet_name='Critical_Swing_Booths', index=False)
    locality_dominance_summary.to_excel(writer, sheet_name='Locality_Dominance', index=False)
    independent_vote_splitting.to_excel(writer, sheet_name='Independent_Splitting', index=False)

wb = load_workbook(excel_filename)
ws_dash = wb['Constituency_Dashboard']
ws_dash.add_image(Image(macro_chart_path), 'E2')

# ==============================================================================
# STEP 7: AUTO-FIT COLUMN WIDTHS ACROSS ALL WORKPASS TABS (FIXED FOR OPENPYXL)
# ==============================================================================
print("Optimizing Excel grid padding layout widths...")
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

wb.save(excel_filename)
print(f"\nSUCCESS! Branded workbook reporting environment finalized: '{excel_filename}'")
